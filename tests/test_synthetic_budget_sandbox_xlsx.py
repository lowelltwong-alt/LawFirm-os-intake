"""Tests for local-only rendering of validated synthetic sandbox drafts."""

from copy import deepcopy
import json
from zipfile import ZipFile

from openpyxl import load_workbook

from lawfirm_os_intake.cli import main
from lawfirm_os_intake.synthetic_budget_sandbox_xlsx import (
    BUDGET_PROPOSAL_REF,
    SANDBOX_EXPORT_REPORT_FILENAME,
    SANDBOX_EXPORT_WORKBOOK_FILENAME,
    run_synthetic_budget_sandbox_xlsx_export,
)
from lawfirm_os_intake.util import digest_text, load_json


FIXED_GENERATED_AT = "2026-07-14T00:00:00Z"
PACKAGE_REF = "fixtures/synthetic/budget-sandbox/synthetic-epli-hours-delta.change-package.json"
SOURCE_REF = "examples/synthetic/labor-employment/replay-inputs/epli-carrier-clean/legal_budget_proposal.json"


def _package(repo_root):
    return load_json(repo_root / PACKAGE_REF)


def _write_package(path, payload):
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def test_sandbox_package_renders_macro_free_local_candidate_workbook(tmp_path, repo_root):
    source_before = (repo_root / SOURCE_REF).read_text(encoding="utf-8")
    report, run_dir = run_synthetic_budget_sandbox_xlsx_export(
        package_path=repo_root / PACKAGE_REF,
        repo_root=repo_root,
        out_dir=tmp_path / "export",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "synthetic_budget_sandbox_xlsx_ready_for_review"
    assert report["failed_check_count"] == 0
    assert report["draft_total"] == 54990.0
    assert report["source_mutation_performed"] is False
    assert report["external_writes_performed"] is False
    assert (repo_root / SOURCE_REF).read_text(encoding="utf-8") == source_before
    assert load_json(run_dir / SANDBOX_EXPORT_REPORT_FILENAME) == report

    workbook = load_workbook(run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME, data_only=False)
    assert workbook.sheetnames == ["Read Me", "Candidate Input Ledger", "Validation"]
    assert workbook["Read Me"]["B8"].value == 54990.0
    assert workbook["Candidate Input Ledger"]["E4"].value == 8.0
    assert workbook["Candidate Input Ledger"]["I4"].value == 3600.0
    assert workbook["Candidate Input Ledger"]["I14"].value == 54990.0
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    with ZipFile(run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME) as archive:
        assert not any(
            forbidden in name
            for name in archive.namelist()
            for forbidden in ("vbaProject", "externalLinks", "connections")
        )


def test_sandbox_package_blocks_stale_lineage_without_workbook(tmp_path, repo_root):
    payload = deepcopy(_package(repo_root))
    payload["source_budget_proposal_sha256"] = "sha256:stale"
    package_path = tmp_path / "tampered.json"
    _write_package(package_path, payload)

    report, run_dir = run_synthetic_budget_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=repo_root,
        out_dir=tmp_path / "blocked",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "blocked_by_synthetic_budget_sandbox_xlsx"
    assert {check["check_id"] for check in report["checks"] if check["status"] == "failed"} >= {
        "source_proposal_hash_matches",
    }
    assert (run_dir / SANDBOX_EXPORT_REPORT_FILENAME).is_file()
    assert not (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_sandbox_package_blocks_hostile_structure_and_negative_amounts(tmp_path, repo_root):
    payload = deepcopy(_package(repo_root))
    payload["blocked_actions"] = [["budget_submission"]]
    payload["lines"][0].update(
        {
            "hourlyRate": -1.0,
            "estimatedFees": -8.0,
            "lineTotal": -8.0,
        }
    )
    payload["fixed_contingency_amount"] = -100.0
    payload["draft_total"] = 54890.0
    payload["delta"] = 800.0
    package_path = tmp_path / "hostile.json"
    _write_package(package_path, payload)

    report, run_dir = run_synthetic_budget_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=repo_root,
        out_dir=tmp_path / "blocked",
        generated_at=FIXED_GENERATED_AT,
    )

    failed = {check["check_id"] for check in report["checks"] if check["status"] == "failed"}
    assert report["status"] == "blocked_by_synthetic_budget_sandbox_xlsx"
    assert {
        "blocked_actions_complete",
        "nonnegative_candidate_amounts",
        "nonnegative_candidate_totals",
    } <= failed
    assert not (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_sandbox_package_uses_decimal_half_up_currency_rules(tmp_path, repo_root):
    payload = deepcopy(_package(repo_root))
    line = payload["lines"][0]
    line.update(
        {
            "estimatedHours": 1.005,
            "hourlyRate": 1.0,
            "estimatedFees": 1.01,
            "estimatedExpenses": 0.0,
            "lineTotal": 1.01,
        }
    )
    payload["draft_total"] = 51391.01
    payload["delta"] = -2698.99
    package_path = tmp_path / "half-up.json"
    _write_package(package_path, payload)

    report, run_dir = run_synthetic_budget_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=repo_root,
        out_dir=tmp_path / "decimal",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "synthetic_budget_sandbox_xlsx_ready_for_review"
    assert (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).is_file()


def test_sandbox_blocks_non_synthetic_pinned_source(tmp_path, repo_root):
    source = load_json(repo_root / SOURCE_REF)
    source["lines"][0]["rate_is_synthetic"] = False
    source_path = tmp_path / BUDGET_PROPOSAL_REF
    source_path.parent.mkdir(parents=True)
    _write_package(source_path, source)

    payload = deepcopy(_package(repo_root))
    payload["source_budget_proposal_sha256"] = digest_text(source_path.read_text(encoding="utf-8"))
    package_path = tmp_path / "non-synthetic-source.json"
    _write_package(package_path, payload)

    report, run_dir = run_synthetic_budget_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=tmp_path,
        out_dir=tmp_path / "blocked",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "blocked_by_synthetic_budget_sandbox_xlsx"
    assert "source_rates_synthetic" in {
        check["check_id"] for check in report["checks"] if check["status"] == "failed"
    }
    assert not (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_sandbox_blocks_malformed_json_with_report_only(tmp_path, repo_root):
    package_path = tmp_path / "malformed.json"
    package_path.write_text("{ not valid json", encoding="utf-8")

    report, run_dir = run_synthetic_budget_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=repo_root,
        out_dir=tmp_path / "blocked",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "blocked_by_synthetic_budget_sandbox_xlsx"
    assert report["checks"][0]["check_id"] == "package_json_parseable"
    assert not (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_sandbox_xlsx_cli_is_local_and_candidate_only(tmp_path, repo_root, capsys):
    code = main(
        [
            "render-synthetic-budget-sandbox-xlsx",
            "--package",
            str(repo_root / PACKAGE_REF),
            "--repo-root",
            str(repo_root),
            "--out-dir",
            str(tmp_path / "cli"),
            "--generated-at",
            FIXED_GENERATED_AT,
        ]
    )

    output = json.loads(capsys.readouterr().out)
    assert code == 0
    assert output["status"] == "synthetic_budget_sandbox_xlsx_ready_for_review"
    assert output["workbook_written"] is True
    assert output["source_mutation_performed"] is False
    assert output["external_writes_performed"] is False
    assert output["lake_write_performed"] is False
    assert output["sqlite_write_performed"] is False
    assert output["budget_submission_authorized"] is False
    assert output["matter_opening_authorized"] is False
