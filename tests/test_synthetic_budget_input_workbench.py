"""Tests for the pinned synthetic budget-input lineage workbench."""

from copy import deepcopy
import json

from openpyxl import load_workbook
import pytest

from lawfirm_os_intake.cli import main
from lawfirm_os_intake.models import SyntheticBudgetInputWorkbenchReport
from lawfirm_os_intake.synthetic_budget_input_workbench import (
    BUDGET_PROPOSAL_REF,
    SYNTHETIC_BUDGET_INPUT_WORKBENCH_REPORT_FILENAME,
    SYNTHETIC_BUDGET_INPUT_WORKBOOK_FILENAME,
    _build_synthetic_budget_input_workbench_report,
    build_synthetic_budget_input_workbench_report,
    run_synthetic_budget_input_workbench,
)
from lawfirm_os_intake.util import load_json


FIXED_GENERATED_AT = "2026-07-13T00:00:00Z"


def _proposal(repo_root):
    return load_json(repo_root / BUDGET_PROPOSAL_REF)


def _write_proposal(path, payload):
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def test_budget_input_workbench_renders_pinned_synthetic_ledger_and_workbook(tmp_path, repo_root):
    report, run_dir = run_synthetic_budget_input_workbench(
        repo_root=repo_root,
        out_dir=tmp_path / "workbench",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report.status == "synthetic_budget_input_workbench_ready_for_review"
    assert report.budget_proposal_id == "le-budget-epli-carrier-clean.v0_1"
    assert report.line_count == 8
    assert report.subtotal_fees == 49990.0
    assert report.subtotal_expenses == 4100.0
    assert report.total_proposed_budget == 54090.0
    assert report.failed_check_count == 0
    assert report.context_lanes[0].inclusion == "used_for_budget_math"
    assert all(lane.inclusion == "excluded_context_only" for lane in report.context_lanes[1:])
    assert report.not_authorized_for_calibration is True
    assert load_json(
        run_dir / SYNTHETIC_BUDGET_INPUT_WORKBENCH_REPORT_FILENAME
    ) == report.model_dump(mode="json")

    workbook = load_workbook(run_dir / SYNTHETIC_BUDGET_INPUT_WORKBOOK_FILENAME, data_only=False)
    assert workbook.sheetnames == ["Read Me", "Input Ledger", "Context Lanes"]
    assert workbook["Input Ledger"]["A1"].value.startswith("Synthetic candidate-only")
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_ui_budget_input_fixture_is_exact_audited_render(repo_root):
    expected = build_synthetic_budget_input_workbench_report(
        repo_root=repo_root,
        generated_at=FIXED_GENERATED_AT,
    )
    fixture = (
        repo_root
        / "apps/legal-intake-budget/src/fixtures/demo-synthetic-budget-input-workbench-report.json"
    )
    assert load_json(fixture) == expected.model_dump(mode="json")


def test_budget_input_workbench_blocks_fee_and_total_arithmetic_drift(tmp_path, repo_root):
    payload = deepcopy(_proposal(repo_root))
    payload["lines"][0]["estimated_fees"] += 1
    proposal_path = tmp_path / "fee-drift.json"
    _write_proposal(proposal_path, payload)

    report = _build_synthetic_budget_input_workbench_report(
        budget_path=proposal_path,
        repo_root=repo_root,
        generated_at=FIXED_GENERATED_AT,
    )

    assert report.status == "blocked_by_synthetic_budget_input_workbench"
    assert {check.check_id for check in report.checks if check.status == "failed"} >= {
        "line_fee_math_reconciles",
        "proposal_totals_reconcile",
    }


def test_budget_input_workbench_blocks_non_synthetic_rates_and_missing_basis_refs(
    tmp_path, repo_root
):
    payload = deepcopy(_proposal(repo_root))
    payload["lines"][0]["rate_is_synthetic"] = False
    payload["lines"][0]["estimate_basis_refs"] = []
    proposal_path = tmp_path / "non-synthetic-rate.json"
    _write_proposal(proposal_path, payload)

    report = _build_synthetic_budget_input_workbench_report(
        budget_path=proposal_path,
        repo_root=repo_root,
        generated_at=FIXED_GENERATED_AT,
    )

    assert report.status == "blocked_by_synthetic_budget_input_workbench"
    assert {check.check_id for check in report.checks if check.status == "failed"} >= {
        "all_rates_synthetic",
        "estimate_basis_refs_present",
    }


def test_budget_input_workbench_keeps_rate_guideline_actuals_and_benchmark_lanes_excluded(
    repo_root,
):
    report = build_synthetic_budget_input_workbench_report(
        repo_root=repo_root,
        generated_at=FIXED_GENERATED_AT,
    )

    excluded_ids = {
        lane.lane_id for lane in report.context_lanes if lane.inclusion == "excluded_context_only"
    }
    assert excluded_ids == {
        "synthetic_rate_card",
        "synthetic_carrier_guideline",
        "synthetic_actuals",
        "synthetic_benchmark",
    }
    assert report.total_proposed_budget == 54090.0


def test_budget_input_workbench_blocks_missing_declared_context_source(monkeypatch, repo_root):
    import lawfirm_os_intake.synthetic_budget_input_workbench as workbench

    monkeypatch.setattr(
        workbench,
        "CONTEXT_SOURCES",
        (
            (
                "missing_context",
                "Missing synthetic context",
                "config/does-not-exist.yaml",
                "Hostile fixture: declared context source is absent.",
            ),
        ),
    )
    report = workbench.build_synthetic_budget_input_workbench_report(
        repo_root=repo_root,
        generated_at=FIXED_GENERATED_AT,
    )

    assert report.status == "blocked_by_synthetic_budget_input_workbench"
    assert "declared_context_sources_present_and_hashed" in {
        check.check_id for check in report.checks if check.status == "failed"
    }


def test_budget_input_workbench_rejects_tampered_displayed_line_integrity(repo_root):
    payload = build_synthetic_budget_input_workbench_report(
        repo_root=repo_root,
        generated_at=FIXED_GENERATED_AT,
    ).model_dump(mode="json")
    payload["lines"][0]["line_total"] += 1
    with pytest.raises(ValueError, match="line total must reconcile"):
        SyntheticBudgetInputWorkbenchReport.model_validate(payload)


def test_budget_input_workbench_cli_uses_fixed_source_and_no_runtime_writes(
    tmp_path, repo_root, capsys
):
    exit_code = main(
        [
            "build-synthetic-budget-input-workbench",
            "--repo-root",
            str(repo_root),
            "--out-dir",
            str(tmp_path / "workbench"),
            "--generated-at",
            FIXED_GENERATED_AT,
        ]
    )

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["status"] == "synthetic_budget_input_workbench_ready_for_review"
    assert output["line_count"] == 8
    assert output["total_proposed_budget"] == 54090.0
    assert output["workbook_written"] is True
    assert output["external_writes_performed"] is False
    assert output["lake_write_performed"] is False
    assert output["sqlite_write_performed"] is False
    assert output["budget_submission_authorized"] is False
    assert output["matter_opening_authorized"] is False
