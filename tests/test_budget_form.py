"""PR5: render a budget proposal into a UTBMS budget form workbook (.xlsx)."""

import json
import re

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from lawfirm_os_intake.budget_form import (
    AMOUNT_HEADER,
    UTBMS_FORM_ROWS,
    form_amounts,
    render_budget_form,
)
from lawfirm_os_intake.cli import main as cli_main
from lawfirm_os_intake.confirmation import bind_confirmation_to_packet_evidence
from lawfirm_os_intake.models import (
    BudgetFormMappingReport,
    BudgetFormTemplateAuditReport,
    HumanConfirmation,
)
from lawfirm_os_intake.util import load_json
from lawfirm_os_intake.workflow import run_budget, run_preflight

_CODE_RE = re.compile(r"\((L\d{3}|E\d{3})\)")


def _budget(tmp_path, repo_root):
    packet, run_dir = run_preflight(
        repo_root / "examples/synthetic/inbound/carrier-assignment-medmal.json",
        repo_root / "context/synthetic-profiles/insurance-defense.yaml",
        tmp_path / "preflight",
    )
    raw = load_json(
        repo_root
        / "examples/synthetic/confirmations/carrier-assignment-medmal.confirmation-template.json"
    )
    raw["preflight_packet_id"] = packet.packet_id
    confirmation = bind_confirmation_to_packet_evidence(
        packet, HumanConfirmation.model_validate(raw)
    )
    confirmation_path = tmp_path / "human_confirmation.json"
    confirmation_path.write_text(confirmation.model_dump_json(), encoding="utf-8")
    budget, _ = run_budget(
        run_dir / "intake_preflight_packet.json",
        confirmation_path,
        repo_root / "context/synthetic-profiles/insurance-defense.yaml",
        tmp_path / "budget",
    )
    return budget


def _fee(budget, code):
    return next(
        line.estimated_fees for line in budget.lines if line.external_code_candidate == code
    )


def _amounts_by_code(ws):
    amount_col = None
    task_col = 1
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(row=r, column=c).value
            if value == AMOUNT_HEADER:
                amount_col = c
            elif isinstance(value, str) and "Phase / Task" in value:
                task_col = c
        if amount_col is not None:
            break
    out = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=task_col).value
        if isinstance(label, str):
            match = _CODE_RE.search(label)
            if match:
                out[match.group(1)] = ws.cell(row=r, column=amount_col).value
    return out


def _phase_tasks():
    phase_tasks = {}
    current = None
    for code, _, kind in UTBMS_FORM_ROWS:
        if kind == "phase":
            current = code
            phase_tasks[current] = []
        elif current is not None:
            phase_tasks[current].append(code)
    return phase_tasks


def _write_structural_template(repo_root, path, *, broken_phase=None, omit_code=None):
    manifest = json.loads(
        (repo_root / "tests/fixtures/budget_form/sanitized_template_structure.json").read_text(
            encoding="utf-8"
        )
    )
    wb = Workbook()
    ws = wb.active
    ws.title = manifest["sheet_name"]
    ws[manifest["total"]["label_cell"]] = "Total Budgeted ($)"
    ws[manifest["total"]["formula_cell"]] = manifest["total"]["formula"]
    headers = manifest["headers"]
    ws[headers["phase_task"]] = "Phase / Task"
    ws[headers["original_budgeted_amount"]] = "Original Budgeted Amount"
    ws[headers["amount_billed_to_date"]] = "Amount Billed to Date"
    ws[headers["original_budget_remaining"]] = "Original Budget Amount Remaining"
    ws[headers["new_budgeted_amount"]] = "New Budgeted Amount"
    code_rows = {code: int(row) for code, row in manifest["code_rows"].items()}
    labels = {code: label for code, label, _ in UTBMS_FORM_ROWS}
    kinds = {code: kind for code, _, kind in UTBMS_FORM_ROWS}
    for code, row in code_rows.items():
        if code == omit_code:
            continue
        prefix = "" if kinds[code] == "phase" else "     "
        ws.cell(row=row, column=1, value=f"{prefix}({code}) {labels[code]}")
        if kinds[code] == "task":
            ws.cell(row=row, column=13, value=f"=G{row}-J{row}")
    for phase, tasks in _phase_tasks().items():
        row = code_rows[phase]
        if phase == broken_phase:
            continue
        refs = "+".join(f"G{code_rows[task]}" for task in tasks)
        ws.cell(row=row, column=7, value=f"={refs}")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def test_form_amounts_split_fees_and_expenses(tmp_path, repo_root):
    budget = _budget(tmp_path, repo_root)
    amounts = form_amounts(budget)
    l_total = round(sum(v for k, v in amounts.items() if k.startswith("L")), 2)
    e_total = round(sum(v for k, v in amounts.items() if k.startswith("E")), 2)
    assert l_total == budget.subtotal_fees
    assert e_total == budget.subtotal_expenses
    assert amounts["E119"] == 30000.0  # expert vendor (E119) from L340
    assert amounts["L330"] == _fee(budget, "L330")


def test_synthetic_form_has_codes_and_total(tmp_path, repo_root):
    budget = _budget(tmp_path, repo_root)
    out = tmp_path / "form.xlsx"
    render_budget_form(budget, out)
    assert out.exists()
    ws = load_workbook(out).active
    by_code = _amounts_by_code(ws)
    assert by_code["L330"] == _fee(budget, "L330")
    assert by_code["E119"] == 30000.0
    total = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Total Budgeted ($)":
            total = ws.cell(row=r, column=2).value
    # The carrier form has no contingency line: form total is fees + expenses.
    assert total == round(budget.subtotal_fees + budget.subtotal_expenses, 2)


def test_fill_existing_form_writes_into_template(tmp_path, repo_root):
    budget = _budget(tmp_path, repo_root)
    template = tmp_path / "template.xlsx"
    _write_structural_template(repo_root, template)
    filled = tmp_path / "filled.xlsx"
    render_budget_form(budget, filled, template_path=template)
    by_code = _amounts_by_code(load_workbook(filled).active)
    assert by_code["L330"] == _fee(budget, "L330")
    assert by_code["E119"] == 30000.0


def test_budget_form_mapping_report_targets_sanitized_structure(tmp_path, repo_root):
    budget = _budget(tmp_path, repo_root)
    template = _write_structural_template(repo_root, tmp_path / "template.xlsx")
    filled = tmp_path / "filled.xlsx"
    report_path = tmp_path / "budget_form_mapping_report.json"

    render_budget_form(
        budget,
        filled,
        template_path=template,
        mapping_report_out=report_path,
    )

    report = BudgetFormMappingReport.model_validate(load_json(report_path))
    assert report.status == "passed"
    assert report.sheet_name == "Sheet1"
    assert report.task_header_cell == "A15"
    assert report.amount_header_cell == "G15"
    assert report.total_cell == "B11"
    assert report.l_code_total == budget.subtotal_fees
    assert report.e_code_total == budget.subtotal_expenses
    assert report.missing_template_codes == []
    assert report.missing_budget_mappings == []
    assert report.unmapped_budget_amount_codes == []
    assert report.not_authorized_for_client_submission is True
    assert report.external_writes_performed is False
    assert report.non_authoritative is True
    mappings = {mapping.code: mapping for mapping in report.code_mappings}
    assert mappings["L330"].amount_cell == "G53"
    assert mappings["E119"].amount_cell == "G123"
    assert mappings["L330"].amount == _fee(budget, "L330")
    assert mappings["E119"].amount == 30000.0
    assert {check.status for check in report.formula_checks} == {"passed"}
    assert filled.exists()


def test_budget_form_mapping_blocks_missing_phase_formula(tmp_path, repo_root):
    budget = _budget(tmp_path, repo_root)
    template = _write_structural_template(
        repo_root, tmp_path / "template.xlsx", broken_phase="L300"
    )
    filled = tmp_path / "filled.xlsx"
    report_path = tmp_path / "budget_form_mapping_report.json"

    with pytest.raises(ValueError, match="phase_l300_original_budget_formula"):
        render_budget_form(
            budget,
            filled,
            template_path=template,
            mapping_report_out=report_path,
        )

    report = BudgetFormMappingReport.model_validate(load_json(report_path))
    assert report.status == "failed"
    assert any(
        check.check_id == "phase_l300_original_budget_formula" and check.status == "failed"
        for check in report.formula_checks
    )
    assert not filled.exists()


def test_budget_form_mapping_blocks_missing_budget_code_row(tmp_path, repo_root):
    budget = _budget(tmp_path, repo_root)
    template = _write_structural_template(repo_root, tmp_path / "template.xlsx", omit_code="E119")
    filled = tmp_path / "filled.xlsx"
    report_path = tmp_path / "budget_form_mapping_report.json"

    with pytest.raises(ValueError, match="missing_budget_mapping:E119"):
        render_budget_form(
            budget,
            filled,
            template_path=template,
            mapping_report_out=report_path,
        )

    report = BudgetFormMappingReport.model_validate(load_json(report_path))
    assert report.status == "failed"
    assert "E119" in report.missing_template_codes
    assert "E119" in report.missing_budget_mappings
    assert "E119" in report.unmapped_budget_amount_codes
    assert not filled.exists()


def test_budget_form_cli_writes_mapping_report_for_template(tmp_path, repo_root):
    budget = _budget(tmp_path, repo_root)
    budget_path = tmp_path / "legal_budget_proposal.json"
    budget_path.write_text(budget.model_dump_json(), encoding="utf-8")
    template = _write_structural_template(repo_root, tmp_path / "template.xlsx")
    out = tmp_path / "filled.xlsx"
    report_path = tmp_path / "budget_form_mapping_report.json"

    assert (
        cli_main(
            [
                "budget-form",
                "--budget",
                str(budget_path),
                "--out",
                str(out),
                "--template",
                str(template),
                "--mapping-report-out",
                str(report_path),
            ]
        )
        == 0
    )

    report = BudgetFormMappingReport.model_validate(load_json(report_path))
    assert report.status == "passed"
    assert out.exists()


def test_budget_form_audit_cli_passes_structural_template(tmp_path, repo_root):
    template = _write_structural_template(repo_root, tmp_path / "template.xlsx")
    report_path = tmp_path / "budget_form_template_audit_report.json"

    assert (
        cli_main(
            [
                "budget-form-audit",
                "--template",
                str(template),
                "--out",
                str(report_path),
            ]
        )
        == 0
    )

    report = BudgetFormTemplateAuditReport.model_validate(load_json(report_path))
    assert report.status == "passed"
    assert report.template_sha256.startswith("sha256:")
    assert report.task_header_cell == "A15"
    assert report.amount_header_cell == "G15"
    assert report.total_cell == "B11"
    assert report.missing_template_codes == []
    assert report.duplicate_template_codes == []
    assert report.external_writes_performed is False
    assert report.non_authoritative is True
    assert any("Original Budgeted Amount" in item for item in report.checklist_items)
    mappings = {mapping.code: mapping for mapping in report.code_mappings}
    assert mappings["L330"].amount_cell == "G53"
    assert mappings["E119"].amount_cell == "G123"


def test_budget_form_audit_cli_fails_broken_template_without_rendering(tmp_path, repo_root):
    template = _write_structural_template(
        repo_root, tmp_path / "template.xlsx", broken_phase="L400"
    )
    report_path = tmp_path / "budget_form_template_audit_report.json"

    assert (
        cli_main(
            [
                "budget-form-audit",
                "--template",
                str(template),
                "--out",
                str(report_path),
            ]
        )
        == 2
    )

    report = BudgetFormTemplateAuditReport.model_validate(load_json(report_path))
    assert report.status == "failed"
    assert any(
        check.check_id == "phase_l400_original_budget_formula" and check.status == "failed"
        for check in report.formula_checks
    )
