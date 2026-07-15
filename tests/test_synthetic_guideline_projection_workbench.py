"""Hostile and reconciliation checks for the synthetic guideline projection workbench."""

from copy import deepcopy
import json
import shutil
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest

from lawfirm_os_intake.cli import main
from lawfirm_os_intake.models import SyntheticGuidelineProjectionWorkbenchReport
from lawfirm_os_intake.synthetic_guideline_projection_workbench import (
    SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_REPORT_FILENAME,
    SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_WORKBOOK_FILENAME,
    PINNED_SOURCE_MANIFEST_REF,
    SOURCE_REFS,
    build_synthetic_guideline_projection_workbench_report,
    run_synthetic_guideline_projection_workbench,
    _safe_text,
)
from lawfirm_os_intake.util import load_json

FIXED_GENERATED_AT = "2026-07-13T00:00:00Z"


def test_projection_workbench_reconciles_two_synthetic_carriers_and_writes_safe_xlsx(
    tmp_path, repo_root
):
    report, run_dir = run_synthetic_guideline_projection_workbench(
        repo_root=repo_root, out_dir=tmp_path / "workbench", generated_at=FIXED_GENERATED_AT
    )

    assert report.status == "synthetic_guideline_projection_workbench_ready_for_review"
    assert report.proposal_total == 148406.0
    assert len(report.views) == 2
    assert report.failed_check_count == 0
    assert all(view.projection.proposed_total == report.proposal_total for view in report.views)
    assert all(
        view.gross_reductions - view.gross_increases == view.net_delta for view in report.views
    )
    assert all(
        all(requirement.status != "unknown" for requirement in view.preapproval_report.requirements)
        for view in report.views
    )
    assert load_json(
        run_dir / SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_REPORT_FILENAME
    ) == report.model_dump(mode="json")

    workbook_path = run_dir / SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_WORKBOOK_FILENAME
    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == ["Read Me", "A Projection", "B Projection"]
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    with ZipFile(workbook_path) as archive:
        forbidden = ("vbaProject", "externalLinks/", "connections.xml", "embeddings/", "oleObject")
        assert not any(any(marker in name for marker in forbidden) for name in archive.namelist())


def test_projection_workbench_ui_fixture_is_an_exact_deterministic_render(repo_root):
    expected = build_synthetic_guideline_projection_workbench_report(
        repo_root=repo_root, generated_at=FIXED_GENERATED_AT
    )
    fixture = (
        repo_root
        / "apps/legal-intake-budget/src/fixtures/demo-synthetic-guideline-projection-workbench-report.json"
    )
    assert load_json(fixture) == expected.model_dump(mode="json")


def test_projection_workbench_blocks_incomplete_or_unevaluable_thresholds(monkeypatch, repo_root):
    import lawfirm_os_intake.guidelines as guidelines

    original = guidelines.load_carrier_guideline
    malformed = deepcopy(original(repo_root / "config/synthetic-carrier-guideline.yaml"))
    del malformed["carriers"]["synthetic-carrier-a"]["pre_approval_thresholds"][
        "experts_over_count"
    ]
    monkeypatch.setattr(guidelines, "load_carrier_guideline", lambda _path: malformed)

    report = build_synthetic_guideline_projection_workbench_report(
        repo_root=repo_root, generated_at=FIXED_GENERATED_AT
    )

    assert report.status == "blocked_by_synthetic_guideline_projection_workbench"
    assert "thresholds_complete_and_evaluable" in {
        check.check_id for check in report.checks if check.status == "failed"
    }


def test_projection_workbench_blocks_pinned_source_digest_drift(tmp_path, repo_root):
    candidate_root = tmp_path / "candidate"
    for _, _, source_ref in SOURCE_REFS:
        source = repo_root / source_ref
        target = candidate_root / source_ref
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source, target)
    manifest_target = candidate_root / PINNED_SOURCE_MANIFEST_REF
    manifest_target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(repo_root / PINNED_SOURCE_MANIFEST_REF, manifest_target)
    rate_card_path = candidate_root / "config/synthetic-carrier-rate-card.yaml"
    rate_card_path.write_text(rate_card_path.read_text(encoding="utf-8") + "\n", encoding="utf-8")

    report = build_synthetic_guideline_projection_workbench_report(
        repo_root=candidate_root, generated_at=FIXED_GENERATED_AT
    )

    assert report.status == "blocked_by_synthetic_guideline_projection_workbench"
    assert "pinned_synthetic_source_hashes_match" in {
        check.check_id for check in report.checks if check.status == "failed"
    }


def test_projection_workbench_model_rejects_tampered_signed_delta(repo_root):
    payload = build_synthetic_guideline_projection_workbench_report(
        repo_root=repo_root, generated_at=FIXED_GENERATED_AT
    ).model_dump(mode="json")
    payload["views"][0]["net_delta"] += 1

    with pytest.raises(ValueError, match="net delta does not reconcile"):
        SyntheticGuidelineProjectionWorkbenchReport.model_validate(payload)


@pytest.mark.parametrize("value", ["=1+1", "+SUM(A1:A2)", "-1+1", "@SUM(A1:A2)"])
def test_projection_workbench_neutralizes_formula_like_xlsx_text(value):
    assert _safe_text(value) == "'" + value


def test_projection_workbench_cli_is_read_only_and_does_not_authorize_submission(
    tmp_path, repo_root, capsys
):
    exit_code = main(
        [
            "build-synthetic-guideline-projection-workbench",
            "--repo-root",
            str(repo_root),
            "--out-dir",
            str(tmp_path / "workbench"),
            "--generated-at",
            FIXED_GENERATED_AT,
        ]
    )

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["status"] == "synthetic_guideline_projection_workbench_ready_for_review"
    assert output["carrier_count"] == 2
    assert output["workbook_written"] is True
    assert output["external_writes_performed"] is False
    assert output["lake_write_performed"] is False
    assert output["sqlite_write_performed"] is False
    assert output["budget_submission_authorized"] is False
