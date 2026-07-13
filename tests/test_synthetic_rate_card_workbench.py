"""Tests for the synthetic-only editable rate-card workbench."""

from copy import deepcopy
import json

import yaml

from lawfirm_os_intake.cli import main
from lawfirm_os_intake.synthetic_rate_card_workbench import (
    SYNTHETIC_RATE_CARD_WORKBENCH_REPORT_FILENAME,
    SYNTHETIC_RATE_CARD_WORKBOOK_FILENAME,
    build_synthetic_rate_card_workbench_report,
    run_synthetic_rate_card_workbench,
)
from lawfirm_os_intake.util import load_json


FIXED_GENERATED_AT = "2026-07-13T00:00:00Z"


def _card_path(repo_root):
    return repo_root / "config" / "synthetic-carrier-rate-card.yaml"


def _write_card(path, card):
    path.write_text(yaml.safe_dump(card, sort_keys=False), encoding="utf-8")


def _card(repo_root):
    return yaml.safe_load(_card_path(repo_root).read_text(encoding="utf-8"))


def test_rate_card_workbench_renders_audited_synthetic_catalog_and_workbook(tmp_path, repo_root):
    report, run_dir = run_synthetic_rate_card_workbench(
        repo_root=repo_root,
        out_dir=tmp_path / "workbench",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report.status == "synthetic_rate_card_workbench_ready_for_review"
    assert report.row_count == 24
    assert report.carrier_count == 2
    assert report.state_count == 3
    assert report.title_count == 4
    assert report.failed_check_count == 0
    assert report.rate_card_ref == "config/synthetic-carrier-rate-card.yaml"
    assert report.rate_card_sha256.startswith("sha256:")
    assert report.data_origin == "synthetic"
    assert report.real_rate_import_allowed is False
    assert report.external_writes_performed is False
    assert report.lake_write_performed is False
    assert report.sqlite_write_performed is False
    assert report.budget_submission_authorized is False
    assert report.matter_opening_authorized is False

    persisted = load_json(run_dir / SYNTHETIC_RATE_CARD_WORKBENCH_REPORT_FILENAME)
    assert persisted == report.model_dump(mode="json")
    workbook_path = run_dir / SYNTHETIC_RATE_CARD_WORKBOOK_FILENAME
    assert workbook_path.is_file()

    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == ["Read Me", "Rate Card", "State Role Summary", "Change Log"]
    assert all("Synthetic" in str(sheet["A1"].value) for sheet in workbook)
    rate_rows = list(workbook["Rate Card"].iter_rows(min_row=4, values_only=True))
    assert len(rate_rows) == report.row_count
    assert {(row[0], row[3], row[4], row[5]) for row in rate_rows} == {
        (row.carrier_id, row.state, row.title, row.hourly_rate) for row in report.rows
    }
    assert all(cell.data_type != "f" for sheet in workbook for row in sheet for cell in row)


def test_rate_card_workbench_is_stable_for_fixed_source_and_timestamp(tmp_path, repo_root):
    first = build_synthetic_rate_card_workbench_report(
        _card_path(repo_root), repo_root=repo_root, generated_at=FIXED_GENERATED_AT
    )
    second = build_synthetic_rate_card_workbench_report(
        _card_path(repo_root), repo_root=repo_root, generated_at=FIXED_GENERATED_AT
    )

    assert first.model_dump(mode="json") == second.model_dump(mode="json")


def test_ui_rate_card_fixture_is_the_exact_audited_source_render(repo_root):
    expected = build_synthetic_rate_card_workbench_report(
        _card_path(repo_root), repo_root=repo_root, generated_at=FIXED_GENERATED_AT
    )
    fixture_path = (
        repo_root
        / "apps/legal-intake-budget/src/fixtures/demo-synthetic-rate-card-workbench-report.json"
    )

    assert load_json(fixture_path) == expected.model_dump(mode="json")


def test_rate_card_workbench_fails_closed_for_real_data_declaration(tmp_path, repo_root):
    card = _card(repo_root)
    card["contains_real_negotiated_rates"] = True
    path = tmp_path / "real-data-declared.yaml"
    _write_card(path, card)

    report = build_synthetic_rate_card_workbench_report(
        path, repo_root=tmp_path, generated_at=FIXED_GENERATED_AT
    )

    assert report.status == "blocked_by_synthetic_rate_card_workbench"
    assert report.failed_check_count == 1
    assert "synthetic_candidate_declaration" in {
        check.check_id for check in report.checks if check.status == "failed"
    }


def test_rate_card_workbench_fails_closed_for_missing_role_and_does_not_write_xlsx(
    tmp_path, repo_root
):
    card = _card(repo_root)
    del card["carriers"]["synthetic-carrier-a"]["schedule"]["NV"]["paralegal"]
    root = tmp_path / "repo"
    (root / "config").mkdir(parents=True)
    _write_card(root / "config" / "synthetic-carrier-rate-card.yaml", card)

    report, run_dir = run_synthetic_rate_card_workbench(
        repo_root=root,
        out_dir=tmp_path / "workbench",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report.status == "blocked_by_synthetic_rate_card_workbench"
    assert "schedule_roles_complete" in {
        check.check_id for check in report.checks if check.status == "failed"
    }
    assert (run_dir / SYNTHETIC_RATE_CARD_WORKBENCH_REPORT_FILENAME).is_file()
    assert not (run_dir / SYNTHETIC_RATE_CARD_WORKBOOK_FILENAME).exists()


def test_rate_card_workbench_counterfactual_changes_only_the_target_cell_and_summary(
    tmp_path, repo_root
):
    baseline = build_synthetic_rate_card_workbench_report(
        _card_path(repo_root), repo_root=repo_root, generated_at=FIXED_GENERATED_AT
    )
    card = deepcopy(_card(repo_root))
    card["carriers"]["synthetic-carrier-a"]["schedule"]["CA"]["partner"] = 530
    path = tmp_path / "counterfactual.yaml"
    _write_card(path, card)
    changed = build_synthetic_rate_card_workbench_report(
        path, repo_root=tmp_path, generated_at=FIXED_GENERATED_AT
    )

    baseline_rates = {
        (row.carrier_id, row.state, row.title): row.hourly_rate for row in baseline.rows
    }
    changed_rates = {
        (row.carrier_id, row.state, row.title): row.hourly_rate for row in changed.rows
    }
    changed_cells = {key for key in baseline_rates if baseline_rates[key] != changed_rates[key]}
    assert changed_cells == {("synthetic-carrier-a", "CA", "partner")}
    assert changed_rates[("synthetic-carrier-a", "CA", "partner")] == 530.0
    assert baseline.data_origin == changed.data_origin == "synthetic"
    assert baseline.real_rate_import_allowed is changed.real_rate_import_allowed is False


def test_rate_card_workbench_cli_uses_fixed_repo_source_and_returns_boundary_flags(
    tmp_path, repo_root, capsys
):
    exit_code = main(
        [
            "build-synthetic-rate-card-workbench",
            "--repo-root",
            str(repo_root),
            "--out-dir",
            str(tmp_path / "workbench"),
            "--generated-at",
            FIXED_GENERATED_AT,
        ]
    )

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["status"] == "synthetic_rate_card_workbench_ready_for_review"
    assert output["row_count"] == 24
    assert output["workbook_written"] is True
    assert output["real_rate_import_allowed"] is False
    assert output["external_writes_performed"] is False
    assert output["lake_write_performed"] is False
    assert output["sqlite_write_performed"] is False
    assert output["budget_submission_authorized"] is False
    assert output["matter_opening_authorized"] is False
