"""Tests for the synthetic editable budget-configuration inventory."""

import json
import shutil

import yaml

from lawfirm_os_intake.cli import main
from lawfirm_os_intake.synthetic_budget_configuration_workbench import (
    REPORT_FILENAME,
    WORKBOOK_FILENAME,
    SOURCE_SPECS,
    build_synthetic_budget_configuration_workbench_report,
    run_synthetic_budget_configuration_workbench,
)
from lawfirm_os_intake.util import load_json


FIXED_TIME = "2026-07-14T00:00:00Z"


def _copy_sources(repo_root, root):
    for _, source_ref, _ in SOURCE_SPECS:
        destination = root / source_ref
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(repo_root / source_ref, destination)


def test_configuration_workbench_inventories_editable_synthetic_inputs(tmp_path, repo_root):
    report, out_dir = run_synthetic_budget_configuration_workbench(
        repo_root=repo_root, out_dir=tmp_path / "workbench", generated_at=FIXED_TIME
    )

    assert report.status == "synthetic_budget_configuration_workbench_ready_for_review"
    assert report.source_count == 4
    assert report.entry_count == 159
    assert report.failed_check_count == 0
    assert report.entries_by_math_effect["proposal_template_hours"] == 55
    assert report.entries_by_math_effect["guideline_projection_rate_cap"] == 8
    assert all(entry.candidate_only and entry.synthetic_only for entry in report.entries)
    assert all(source.source_sha256.startswith("sha256:") for source in report.sources)
    assert report.real_rate_import_allowed is False
    assert report.configuration_import_performed is False
    assert report.external_writes_performed is False

    assert load_json(out_dir / REPORT_FILENAME) == report.model_dump(mode="json")
    workbook_path = out_dir / WORKBOOK_FILENAME
    assert workbook_path.is_file()
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == ["Read Me", "Editable Values", "Source Manifest"]
    assert len(list(workbook["Editable Values"].iter_rows(min_row=4, values_only=True))) == 159
    assert all(cell.data_type != "f" for sheet in workbook for row in sheet for cell in row)


def test_configuration_workbench_is_deterministic_for_fixed_sources(repo_root):
    first = build_synthetic_budget_configuration_workbench_report(
        repo_root=repo_root, generated_at=FIXED_TIME
    )
    second = build_synthetic_budget_configuration_workbench_report(
        repo_root=repo_root, generated_at=FIXED_TIME
    )

    assert first.model_dump(mode="json") == second.model_dump(mode="json")


def test_configuration_workbench_fails_closed_for_real_rate_declaration(tmp_path, repo_root):
    root = tmp_path / "repo"
    _copy_sources(repo_root, root)
    rate_card_path = root / "config/synthetic-carrier-rate-card.yaml"
    rate_card = yaml.safe_load(rate_card_path.read_text(encoding="utf-8"))
    rate_card["contains_real_negotiated_rates"] = True
    rate_card_path.write_text(yaml.safe_dump(rate_card, sort_keys=False), encoding="utf-8")

    report, out_dir = run_synthetic_budget_configuration_workbench(
        repo_root=root, out_dir=tmp_path / "blocked", generated_at=FIXED_TIME
    )

    assert report.status == "blocked_by_synthetic_budget_configuration_workbench"
    assert "synthetic_declarations_hold" in {
        check.check_id for check in report.checks if check.status == "failed"
    }
    assert not (out_dir / WORKBOOK_FILENAME).exists()


def test_configuration_workbench_cli_reports_no_import_or_external_write(
    tmp_path, repo_root, capsys
):
    code = main(
        [
            "build-synthetic-budget-configuration-workbench",
            "--repo-root",
            str(repo_root),
            "--out-dir",
            str(tmp_path),
            "--generated-at",
            FIXED_TIME,
        ]
    )
    payload = json.loads(capsys.readouterr().out)

    assert code == 0
    assert payload["entry_count"] == 159
    assert payload["configuration_import_performed"] is False
    assert payload["external_writes_performed"] is False
    assert payload["lake_write_performed"] is False
    assert payload["sqlite_write_performed"] is False
