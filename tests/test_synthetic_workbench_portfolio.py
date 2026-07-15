"""Independent portfolio acceptance check for the synthetic budget workbenches."""

import shutil
from zipfile import ZipFile

import yaml
from openpyxl import load_workbook

from lawfirm_os_intake.synthetic_actuals_workbench import build_synthetic_actuals_workbench_report
from lawfirm_os_intake.synthetic_budget_configuration_change import (
    build_synthetic_budget_configuration_change_package,
)
from lawfirm_os_intake.synthetic_budget_configuration_workbench import (
    SOURCE_SPECS,
    build_synthetic_budget_configuration_workbench_report,
)
from lawfirm_os_intake.synthetic_budget_input_workbench import (
    build_synthetic_budget_input_workbench_report,
)
from lawfirm_os_intake.synthetic_budget_sandbox_xlsx import (
    SANDBOX_EXPORT_REPORT_FILENAME as BUDGET_SANDBOX_REPORT_FILENAME,
    SANDBOX_EXPORT_WORKBOOK_FILENAME as BUDGET_SANDBOX_WORKBOOK_FILENAME,
    run_synthetic_budget_sandbox_xlsx_export,
)
from lawfirm_os_intake.synthetic_configuration_regeneration_binding import (
    build_synthetic_configuration_regeneration_binding_report,
)
from lawfirm_os_intake.synthetic_guideline_projection_workbench import (
    SOURCE_REFS,
    build_synthetic_guideline_projection_workbench_report,
)
from lawfirm_os_intake.synthetic_rate_card_workbench import (
    build_synthetic_rate_card_workbench_report,
)
from lawfirm_os_intake.synthetic_rate_card_sandbox_xlsx import (
    SANDBOX_EXPORT_REPORT_FILENAME as RATE_CARD_SANDBOX_REPORT_FILENAME,
    SANDBOX_EXPORT_WORKBOOK_FILENAME as RATE_CARD_SANDBOX_WORKBOOK_FILENAME,
    run_synthetic_rate_card_sandbox_xlsx_export,
)
from lawfirm_os_intake.synthetic_rejection_appeal_workbench import (
    build_synthetic_rejection_appeal_workbench_report,
)


FIXED_TIME = "2026-07-14T00:00:00Z"


def _candidate_root(repo_root, root):
    refs = {ref for _, ref, _ in SOURCE_SPECS} | {ref for _, _, ref in SOURCE_REFS}
    for source_ref in refs:
        target = root / source_ref
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(repo_root / source_ref, target)
    rate_path = root / "config/synthetic-carrier-rate-card.yaml"
    rate_card = yaml.safe_load(rate_path.read_text(encoding="utf-8"))
    rate_card["carriers"]["synthetic-carrier-a"]["schedule"]["NV"]["partner"] = 455
    rate_path.write_text(yaml.safe_dump(rate_card, sort_keys=False), encoding="utf-8")


def test_synthetic_budget_workbench_portfolio_acceptance(tmp_path, repo_root):
    candidate_root = tmp_path / "candidate"
    _candidate_root(repo_root, candidate_root)
    budget_source_ref = (
        "examples/synthetic/labor-employment/replay-inputs/epli-carrier-clean/"
        "legal_budget_proposal.json"
    )
    rate_card_source_ref = "config/synthetic-carrier-rate-card.yaml"
    pinned_refs = {
        budget_source_ref: (repo_root / budget_source_ref).read_text(encoding="utf-8"),
        rate_card_source_ref: (repo_root / rate_card_source_ref).read_text(encoding="utf-8"),
        "fixtures/synthetic/budget-sandbox/synthetic-epli-hours-delta.change-package.json": (
            repo_root / "fixtures/synthetic/budget-sandbox/"
            "synthetic-epli-hours-delta.change-package.json"
        ).read_text(encoding="utf-8"),
        "fixtures/synthetic/rate-card-sandbox/"
        "synthetic-rate-card-nv-partner-delta.change-package.json": (
            repo_root / "fixtures/synthetic/rate-card-sandbox/"
            "synthetic-rate-card-nv-partner-delta.change-package.json"
        ).read_text(encoding="utf-8"),
    }
    reports = [
        build_synthetic_rate_card_workbench_report(
            repo_root / "config/synthetic-carrier-rate-card.yaml",
            repo_root=repo_root,
            generated_at=FIXED_TIME,
        ),
        build_synthetic_budget_input_workbench_report(repo_root=repo_root, generated_at=FIXED_TIME),
        build_synthetic_guideline_projection_workbench_report(
            repo_root=repo_root, generated_at=FIXED_TIME
        ),
        build_synthetic_actuals_workbench_report(repo_root=repo_root, generated_at=FIXED_TIME),
        build_synthetic_rejection_appeal_workbench_report(
            repo_root=repo_root, generated_at=FIXED_TIME
        ),
        build_synthetic_budget_configuration_workbench_report(
            repo_root=repo_root, generated_at=FIXED_TIME
        ),
        build_synthetic_budget_configuration_change_package(
            baseline_root=repo_root, candidate_root=candidate_root, generated_at=FIXED_TIME
        ),
        build_synthetic_configuration_regeneration_binding_report(
            baseline_root=repo_root, candidate_root=candidate_root, generated_at=FIXED_TIME
        ),
    ]
    budget_sandbox_report, budget_sandbox_dir = run_synthetic_budget_sandbox_xlsx_export(
        package_path=(
            repo_root / "fixtures/synthetic/budget-sandbox/"
            "synthetic-epli-hours-delta.change-package.json"
        ),
        repo_root=repo_root,
        out_dir=tmp_path / "budget-sandbox",
        generated_at=FIXED_TIME,
    )
    rate_card_sandbox_report, rate_card_sandbox_dir = run_synthetic_rate_card_sandbox_xlsx_export(
        package_path=(
            repo_root / "fixtures/synthetic/rate-card-sandbox/"
            "synthetic-rate-card-nv-partner-delta.change-package.json"
        ),
        repo_root=repo_root,
        out_dir=tmp_path / "rate-card-sandbox",
        generated_at=FIXED_TIME,
    )
    assert all("ready" in report.status for report in reports)
    required_boundary_fields = {
        "candidate_only": True,
        "external_writes_performed": False,
        "lake_write_performed": False,
        "sqlite_write_performed": False,
        "budget_submission_authorized": False,
        "matter_opening_authorized": False,
        "silent_learning_performed": False,
    }
    for report in reports:
        for field, expected in required_boundary_fields.items():
            assert hasattr(report, field), f"{type(report).__name__} must declare {field}"
            assert getattr(report, field) is expected
    assert reports[-2].changed_source_ids == ["rate_card"]
    assert reports[-1].changed_source_ids == ["rate_card"]
    assert budget_sandbox_report["status"] == "synthetic_budget_sandbox_xlsx_ready_for_review"
    assert budget_sandbox_report["draft_total"] == 54990.0
    assert budget_sandbox_report["delta"] == 900.0
    assert sorted(path.name for path in budget_sandbox_dir.iterdir()) == sorted(
        [BUDGET_SANDBOX_REPORT_FILENAME, BUDGET_SANDBOX_WORKBOOK_FILENAME]
    )
    budget_workbook_path = budget_sandbox_dir / BUDGET_SANDBOX_WORKBOOK_FILENAME
    budget_workbook = load_workbook(budget_workbook_path, data_only=False)
    assert budget_workbook.sheetnames == ["Read Me", "Candidate Input Ledger", "Validation"]
    assert budget_workbook["Read Me"]["B8"].value == 54990.0
    assert rate_card_sandbox_report["status"] == "synthetic_rate_card_sandbox_xlsx_ready_for_review"
    assert rate_card_sandbox_report["draft_rate_total"] == 6995.0
    assert rate_card_sandbox_report["delta"] == 5.0
    assert rate_card_sandbox_report["rate_card_applied_to_budget"] is False
    assert sorted(path.name for path in rate_card_sandbox_dir.iterdir()) == sorted(
        [RATE_CARD_SANDBOX_REPORT_FILENAME, RATE_CARD_SANDBOX_WORKBOOK_FILENAME]
    )
    rate_card_workbook_path = rate_card_sandbox_dir / RATE_CARD_SANDBOX_WORKBOOK_FILENAME
    rate_card_workbook = load_workbook(rate_card_workbook_path, data_only=False)
    assert rate_card_workbook.sheetnames == [
        "Read Me",
        "Candidate Rate Card",
        "Candidate State Summary",
        "Validation",
    ]
    assert rate_card_workbook["Read Me"]["B8"].value == 6995.0
    for workbook_path in (budget_workbook_path, rate_card_workbook_path):
        workbook = load_workbook(workbook_path, data_only=False)
        assert not any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        with ZipFile(workbook_path) as archive:
            assert not any(
                forbidden in name
                for name in archive.namelist()
                for forbidden in ("vbaProject", "externalLinks", "connections")
            )
    for report in (budget_sandbox_report, rate_card_sandbox_report):
        assert report["candidate_only"] is True
        assert report["non_authoritative"] is True
        assert report["local_output_only"] is True
        assert report["source_mutation_performed"] is False
        assert report["external_writes_performed"] is False
        assert report["lake_write_performed"] is False
        assert report["sqlite_write_performed"] is False
        assert report["budget_submission_authorized"] is False
        assert report["matter_opening_authorized"] is False
        assert report["silent_learning_performed"] is False
    for source_ref, expected in pinned_refs.items():
        assert (repo_root / source_ref).read_text(encoding="utf-8") == expected
