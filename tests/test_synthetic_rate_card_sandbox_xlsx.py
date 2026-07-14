"""Tests for local-only rendering of validated synthetic rate-card drafts."""

from copy import deepcopy
import json
from zipfile import ZipFile

import yaml
from openpyxl import load_workbook

from lawfirm_os_intake.cli import main
from lawfirm_os_intake.synthetic_rate_card_sandbox_xlsx import (
    RATE_CARD_REF,
    SANDBOX_EXPORT_REPORT_FILENAME,
    SANDBOX_EXPORT_WORKBOOK_FILENAME,
    run_synthetic_rate_card_sandbox_xlsx_export,
)
from lawfirm_os_intake.util import digest_text, load_json


FIXED_GENERATED_AT = "2026-07-14T00:00:00Z"
PACKAGE_REF = (
    "fixtures/synthetic/rate-card-sandbox/synthetic-rate-card-nv-partner-delta.change-package.json"
)


def _package(repo_root):
    return load_json(repo_root / PACKAGE_REF)


def _write_json(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def _copy_source(root, repo_root):
    source_path = root / RATE_CARD_REF
    source_path.parent.mkdir(parents=True)
    source_path.write_text(
        (repo_root / RATE_CARD_REF).read_text(encoding="utf-8"), encoding="utf-8"
    )
    return source_path


def test_rate_card_sandbox_renders_macro_free_candidate_workbook(tmp_path, repo_root):
    source_before = (repo_root / RATE_CARD_REF).read_text(encoding="utf-8")
    report, run_dir = run_synthetic_rate_card_sandbox_xlsx_export(
        package_path=repo_root / PACKAGE_REF,
        repo_root=repo_root,
        out_dir=tmp_path / "export",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "synthetic_rate_card_sandbox_xlsx_ready_for_review"
    assert report["failed_check_count"] == 0
    assert report["pinned_rate_total"] == 6990.0
    assert report["draft_rate_total"] == 6995.0
    assert report["delta"] == 5.0
    assert report["changed_cell_count"] == 1
    assert report["rate_card_applied_to_budget"] is False
    assert report["source_mutation_performed"] is False
    assert (repo_root / RATE_CARD_REF).read_text(encoding="utf-8") == source_before
    assert load_json(run_dir / SANDBOX_EXPORT_REPORT_FILENAME) == report

    workbook_path = run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME
    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook.sheetnames == [
        "Read Me",
        "Candidate Rate Card",
        "Candidate State Summary",
        "Validation",
    ]
    assert workbook["Read Me"]["B8"].value == 6995.0
    assert workbook["Candidate Rate Card"]["B10"].value == "Harbor Point Insurance"
    assert workbook["Candidate Rate Card"]["F10"].value == 455.0
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    with ZipFile(workbook_path) as archive:
        assert not any(
            forbidden in name
            for name in archive.namelist()
            for forbidden in ("vbaProject", "externalLinks", "connections")
        )


def test_rate_card_sandbox_blocks_stale_lineage_and_hostile_values(tmp_path, repo_root):
    payload = deepcopy(_package(repo_root))
    payload["source_rate_card_sha256"] = "sha256:stale"
    payload["blocked_actions"] = [["budget_submission"]]
    payload["cells"][0]["hourlyRate"] = -1.0
    package_path = tmp_path / "hostile.json"
    _write_json(package_path, payload)

    report, run_dir = run_synthetic_rate_card_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=repo_root,
        out_dir=tmp_path / "blocked",
        generated_at=FIXED_GENERATED_AT,
    )

    failed = {check["check_id"] for check in report["checks"] if check["status"] == "failed"}
    assert report["status"] == "blocked_by_synthetic_rate_card_sandbox_xlsx"
    assert {
        "source_rate_card_hash_matches",
        "blocked_actions_complete",
        "positive_cent_precision_rates",
    } <= failed
    assert not (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_rate_card_sandbox_blocks_source_with_real_rate_declaration(tmp_path, repo_root):
    source_path = _copy_source(tmp_path / "repo", repo_root)
    card = yaml.safe_load(source_path.read_text(encoding="utf-8"))
    card["contains_real_negotiated_rates"] = True
    source_path.write_text(yaml.safe_dump(card, sort_keys=False), encoding="utf-8")

    payload = deepcopy(_package(repo_root))
    payload["source_rate_card_sha256"] = digest_text(source_path.read_text(encoding="utf-8"))
    package_path = tmp_path / "matching-source.json"
    _write_json(package_path, payload)

    report, run_dir = run_synthetic_rate_card_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=tmp_path / "repo",
        out_dir=tmp_path / "blocked",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "blocked_by_synthetic_rate_card_sandbox_xlsx"
    assert "source_rate_card_ready" in {
        check["check_id"] for check in report["checks"] if check["status"] == "failed"
    }
    assert not (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_rate_card_sandbox_blocks_malformed_json_with_report_only(tmp_path, repo_root):
    package_path = tmp_path / "malformed.json"
    package_path.write_text("{ invalid json", encoding="utf-8")

    report, run_dir = run_synthetic_rate_card_sandbox_xlsx_export(
        package_path=package_path,
        repo_root=repo_root,
        out_dir=tmp_path / "blocked",
        generated_at=FIXED_GENERATED_AT,
    )

    assert report["status"] == "blocked_by_synthetic_rate_card_sandbox_xlsx"
    assert [check["check_id"] for check in report["checks"][:2]] == [
        "source_rate_card_readable",
        "package_json_parseable",
    ]
    assert not (run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_rate_card_sandbox_reports_missing_source_without_workbook(tmp_path, repo_root, capsys):
    package_path = repo_root / PACKAGE_REF

    code = main(
        [
            "render-synthetic-rate-card-sandbox-xlsx",
            "--package",
            str(package_path),
            "--repo-root",
            str(tmp_path / "missing-repo"),
            "--out-dir",
            str(tmp_path / "blocked"),
            "--generated-at",
            FIXED_GENERATED_AT,
        ]
    )

    output = json.loads(capsys.readouterr().out)
    report = load_json(tmp_path / "blocked" / SANDBOX_EXPORT_REPORT_FILENAME)
    assert code == 2
    assert output["status"] == "blocked_by_synthetic_rate_card_sandbox_xlsx"
    assert report["checks"][0]["check_id"] == "source_rate_card_readable"
    assert report["checks"][0]["status"] == "failed"
    assert not (tmp_path / "blocked" / SANDBOX_EXPORT_WORKBOOK_FILENAME).exists()


def test_rate_card_sandbox_cli_is_local_and_candidate_only(tmp_path, repo_root, capsys):
    code = main(
        [
            "render-synthetic-rate-card-sandbox-xlsx",
            "--package",
            str(repo_root / PACKAGE_REF),
            "--repo-root",
            str(repo_root),
            "--out-dir",
            str(tmp_path / "cli"),
            "--generated-at",
            FIXED_GENERATED_AT,
        ]
    )

    output = json.loads(capsys.readouterr().out)
    assert code == 0
    assert output["status"] == "synthetic_rate_card_sandbox_xlsx_ready_for_review"
    assert output["workbook_written"] is True
    assert output["rate_card_applied_to_budget"] is False
    assert output["source_mutation_performed"] is False
    assert output["external_writes_performed"] is False
    assert output["lake_write_performed"] is False
    assert output["sqlite_write_performed"] is False
    assert output["budget_submission_authorized"] is False
    assert output["matter_opening_authorized"] is False
