"""Render validated browser rate-card drafts as local synthetic XLSX artifacts.

The browser draft is intentionally untrusted. This module validates a complete
candidate rate-card package against the fixed synthetic catalog before writing
local review evidence. It never changes the rate-card source or applies a rate
to a budget.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import yaml

from .synthetic_rate_card_workbench import build_synthetic_rate_card_workbench_report
from .util import digest_json, digest_text, now_iso, write_json


RATE_CARD_REF = "config/synthetic-carrier-rate-card.yaml"
SANDBOX_EXPORT_REPORT_FILENAME = "synthetic_rate_card_sandbox_xlsx_export_report.json"
SANDBOX_EXPORT_WORKBOOK_FILENAME = "synthetic_rate_card_sandbox_candidate.xlsx"
METHODOLOGY_VERSION = "synthetic_rate_card_sandbox_xlsx_export.v0_1"
REQUIRED_BLOCKED_ACTIONS = {
    "configuration_write",
    "real_rate_import",
    "rate_card_apply_to_budget",
    "budget_submission",
    "exception_lake_write",
    "sqlite_write",
    "silent_learning",
}
CENT = Decimal("0.01")


def _decimal(value: Any) -> Decimal:
    """Read an untrusted positive rate or total at exact cent precision."""
    if isinstance(value, bool):
        raise ValueError("Boolean values are not valid rate-card amounts.")
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Rate-card values must be finite decimal numbers.") from exc
    if not parsed.is_finite():
        raise ValueError("Rate-card values must be finite decimal numbers.")
    rounded = parsed.quantize(CENT)
    if parsed != rounded:
        raise ValueError("Rate-card amounts must already have cent precision.")
    return rounded


def _check(check_id: str, passed: bool, message: str, *refs: str) -> dict[str, Any]:
    return {
        "check_id": check_id,
        "status": "passed" if passed else "failed",
        "message": message,
        "evidence_refs": [ref for ref in refs if ref],
    }


def _source_cells(report: Any) -> list[dict[str, Any]]:
    return [
        {
            "cellId": f"{row.carrier_id}|{row.state}|{row.title}",
            "carrierId": row.carrier_id,
            "carrierName": row.carrier_name,
            "effectiveDate": row.effective_date,
            "state": row.state,
            "title": row.title,
            "hourlyRate": _decimal(row.hourly_rate),
        }
        for row in report.rows
    ]


def _validate_package(
    package: Any,
    *,
    source_hash: str,
    source_ready: bool,
    source_cells: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not isinstance(package, dict):
        return [_check("package_is_mapping", False, "Candidate package must be a JSON object.")]

    blocked_actions = package.get("blocked_actions")
    blocked_actions_valid = isinstance(blocked_actions, list) and all(
        isinstance(action, str) for action in blocked_actions
    )
    checks = [
        _check(
            "source_rate_card_ready",
            source_ready,
            "The pinned source rate card must remain an audited synthetic candidate catalog.",
            RATE_CARD_REF,
        ),
        _check(
            "candidate_synthetic_declaration",
            package.get("artifact_type") == "synthetic_rate_card_sandbox_change_package"
            and package.get("data_origin") == "synthetic"
            and package.get("candidate_only") is True
            and package.get("local_browser_draft") is True,
            "Package must declare the synthetic browser-draft candidate contract.",
        ),
        _check(
            "source_rate_card_hash_matches",
            package.get("source_rate_card_sha256") == source_hash,
            "Package must bind to the current pinned synthetic rate-card hash.",
            RATE_CARD_REF,
        ),
        _check(
            "blocked_actions_complete",
            blocked_actions_valid and REQUIRED_BLOCKED_ACTIONS <= set(blocked_actions),
            "Package must explicitly retain every no-write and no-application boundary.",
        ),
    ]

    cells = package.get("cells")
    checks.append(
        _check(
            "cell_count_matches_source",
            isinstance(cells, list) and len(cells) == len(source_cells),
            "Candidate package must contain every pinned rate-card cell exactly once.",
            RATE_CARD_REF,
        )
    )
    if not isinstance(cells, list):
        return checks

    source_ids = [cell["cellId"] for cell in source_cells]
    seen: set[str] = set()
    identity_ok = len(cells) == len(source_cells)
    rate_values_ok = True
    changed_count = 0
    candidate_total = Decimal("0")
    for index, cell in enumerate(cells):
        if not isinstance(cell, dict):
            identity_ok = False
            rate_values_ok = False
            continue
        cell_id = cell.get("cellId")
        expected = source_cells[index] if index < len(source_cells) else None
        if (
            not isinstance(cell_id, str)
            or cell_id in seen
            or expected is None
            or cell_id != source_ids[index]
            or cell.get("carrierId") != expected["carrierId"]
            or cell.get("state") != expected["state"]
            or cell.get("title") != expected["title"]
        ):
            identity_ok = False
        else:
            seen.add(cell_id)
        try:
            rate = _decimal(cell["hourlyRate"])
            if rate <= 0:
                rate_values_ok = False
            else:
                candidate_total += rate
                if expected is not None and rate != expected["hourlyRate"]:
                    changed_count += 1
        except (KeyError, TypeError, ValueError, InvalidOperation):
            rate_values_ok = False

    checks.extend(
        [
            _check(
                "cell_identity_complete",
                identity_ok and len(seen) == len(source_cells),
                "Candidate cell IDs, carrier IDs, states, and titles must preserve pinned order.",
                RATE_CARD_REF,
            ),
            _check(
                "positive_cent_precision_rates",
                rate_values_ok,
                "Every candidate hourly rate must be a positive finite cent amount.",
            ),
        ]
    )

    try:
        pinned_total = _decimal(package["pinnedRateTotal"])
        draft_total = _decimal(package["draftRateTotal"])
        delta = _decimal(package["delta"])
        package_changed_count = package["changedCellCount"]
        source_total = sum((cell["hourlyRate"] for cell in source_cells), start=Decimal("0"))
        checks.extend(
            [
                _check(
                    "pinned_total_matches_source",
                    pinned_total == source_total,
                    "Pinned rate total must equal the current synthetic rate-card total.",
                    RATE_CARD_REF,
                ),
                _check(
                    "draft_total_reconciles",
                    draft_total == candidate_total,
                    "Draft rate total must equal the sum of candidate cells.",
                ),
                _check(
                    "delta_reconciles",
                    delta == draft_total - pinned_total,
                    "Candidate delta must equal draft total minus pinned total.",
                ),
                _check(
                    "changed_cell_count_reconciles",
                    isinstance(package_changed_count, int)
                    and not isinstance(package_changed_count, bool)
                    and package_changed_count == changed_count,
                    "Changed-cell count must exactly match candidate cells differing from source.",
                ),
            ]
        )
    except (KeyError, TypeError, ValueError, InvalidOperation):
        checks.extend(
            [
                _check(
                    "pinned_total_matches_source", False, "Package totals are missing or invalid."
                ),
                _check("draft_total_reconciles", False, "Package totals are missing or invalid."),
                _check("delta_reconciles", False, "Package totals are missing or invalid."),
                _check(
                    "changed_cell_count_reconciles",
                    False,
                    "Package changed-cell count is missing or invalid.",
                ),
            ]
        )
    return checks


def _write_workbook(
    path: Path,
    *,
    package: dict[str, Any],
    source_hash: str,
    package_hash: str,
    source_cells: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    read_me = workbook.active
    read_me.title = "Read Me"
    read_me.append(["Synthetic Rate Card Sandbox Candidate Workbook"])
    read_me.append([])
    for row in [
        ["Status", "candidate_only_local_export"],
        ["Source rate card", RATE_CARD_REF],
        ["Source rate card hash", source_hash],
        ["Candidate package hash", package_hash],
        ["Pinned rate total", package["pinnedRateTotal"]],
        ["Candidate rate total", package["draftRateTotal"]],
        ["Candidate delta", package["delta"]],
        ["Changed cells", package["changedCellCount"]],
        [
            "Boundary",
            "No source configuration write, real-rate import, rate-card application, submission, Lake/SQLite write, or silent learning.",
        ],
    ]:
        read_me.append(row)
    read_me["A1"].font = Font(bold=True, size=14)
    for cell in read_me[3]:
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    read_me.column_dimensions["A"].width = 26
    read_me.column_dimensions["B"].width = 124

    card = workbook.create_sheet("Candidate Rate Card")
    card.append(
        ["Validated synthetic browser draft. Candidate-only; values are not rate authority."]
    )
    card.append([])
    card.append(
        ["Carrier ID", "Carrier", "Effective Date", "State", "Title", "Candidate Hourly Rate"]
    )
    source_by_id = {cell["cellId"]: cell for cell in source_cells}
    for cell in package["cells"]:
        source = source_by_id[cell["cellId"]]
        card.append(
            [
                cell["carrierId"],
                source["carrierName"],
                source["effectiveDate"],
                cell["state"],
                cell["title"],
                cell["hourlyRate"],
            ]
        )
    for heading in card[3]:
        heading.fill = PatternFill("solid", fgColor="D9EAD3")
        heading.font = Font(bold=True)
    card.freeze_panes = "A4"
    card.auto_filter.ref = f"A3:F{card.max_row}"
    for column, width in {
        "A": 26,
        "B": 28,
        "C": 16,
        "D": 12,
        "E": 24,
        "F": 24,
    }.items():
        card.column_dimensions[column].width = width
    for cell in card["F"][3:]:
        cell.number_format = "$#,##0.00"

    summaries: dict[tuple[str, str], list[Decimal]] = {}
    carrier_names = {cell["carrierId"]: cell["carrierName"] for cell in source_cells}
    for cell in package["cells"]:
        key = (str(cell["carrierId"]), str(cell["state"]))
        summaries.setdefault(key, []).append(_decimal(cell["hourlyRate"]))
    summary = workbook.create_sheet("Candidate State Summary")
    summary.append(["Candidate-only summary. These cells are not a billing or carrier authority."])
    summary.append([])
    summary.append(["Carrier ID", "Carrier", "State", "Roles", "Minimum", "Maximum", "Average"])
    for (carrier_id, state), rates in sorted(summaries.items()):
        summary.append(
            [
                carrier_id,
                carrier_names[carrier_id],
                state,
                len(rates),
                float(min(rates)),
                float(max(rates)),
                float(sum(rates) / len(rates)),
            ]
        )
    for heading in summary[3]:
        heading.fill = PatternFill("solid", fgColor="D9EAD3")
        heading.font = Font(bold=True)
    summary.freeze_panes = "A4"
    summary.auto_filter.ref = f"A3:G{summary.max_row}"
    for column, width in {
        "A": 26,
        "B": 28,
        "C": 12,
        "D": 12,
        "E": 16,
        "F": 16,
        "G": 16,
    }.items():
        summary.column_dimensions[column].width = width
    for column in ("E", "F", "G"):
        for cell in summary[column][3:]:
            cell.number_format = "$#,##0.00"

    validation = workbook.create_sheet("Validation")
    validation.append(["Validation checks are recorded in the companion JSON report."])
    validation.append([])
    validation.append(["Boundary", "Value"])
    for row in [
        ["Data origin", "synthetic"],
        ["Candidate only", True],
        ["Rate-card application", False],
        ["External writes", False],
        ["Budget submission authorized", False],
    ]:
        validation.append(row)
    for heading in validation[3]:
        heading.fill = PatternFill("solid", fgColor="D9EAD3")
        heading.font = Font(bold=True)
    validation.column_dimensions["A"].width = 34
    validation.column_dimensions["B"].width = 88

    workbook.properties.title = "Synthetic Rate Card Sandbox Candidate"
    workbook.properties.subject = "Local candidate-only synthetic evidence"
    workbook.properties.creator = "LawFirm-os-intake"
    workbook.properties.lastModifiedBy = "LawFirm-os-intake"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def run_synthetic_rate_card_sandbox_xlsx_export(
    *,
    package_path: str | Path,
    repo_root: str | Path,
    out_dir: str | Path,
    generated_at: str | None = None,
) -> tuple[dict[str, Any], Path]:
    """Validate a synthetic candidate package and write local review-only evidence."""
    root = Path(repo_root)
    source_path = root / RATE_CARD_REF
    try:
        source_report = build_synthetic_rate_card_workbench_report(
            source_path, repo_root=root, generated_at=generated_at
        )
        source_hash = digest_text(source_path.read_text(encoding="utf-8"))
        source_ready = (
            source_report.status == "synthetic_rate_card_workbench_ready_for_review"
            and source_report.data_origin == "synthetic"
            and source_report.candidate_only
            and source_report.synthetic_only
            and not source_report.real_rate_import_allowed
        )
        source_cells = _source_cells(source_report)
        source_readable_check = _check(
            "source_rate_card_readable",
            True,
            "The pinned synthetic rate card can be read and audited.",
            RATE_CARD_REF,
        )
    except (OSError, ValueError, TypeError, yaml.YAMLError) as exc:
        source_hash = digest_text("")
        source_ready = False
        source_cells = []
        source_readable_check = _check(
            "source_rate_card_readable",
            False,
            f"Pinned source rate card cannot be read and audited: {exc}",
            RATE_CARD_REF,
        )

    package_file = Path(package_path)
    try:
        package_text = package_file.read_text(encoding="utf-8")
        package = json.loads(package_text)
        package_parse_check: dict[str, Any] | None = None
    except (OSError, json.JSONDecodeError) as exc:
        package_text = ""
        package = None
        package_parse_check = _check(
            "package_json_parseable", False, f"Candidate package cannot be read as JSON: {exc}"
        )
    package_hash = digest_text(package_text)
    checks = _validate_package(
        package,
        source_hash=source_hash,
        source_ready=source_ready,
        source_cells=source_cells,
    )
    checks.insert(0, source_readable_check)
    if package_parse_check is not None:
        checks.insert(1, package_parse_check)
    failed = [check for check in checks if check["status"] == "failed"]
    status = (
        "synthetic_rate_card_sandbox_xlsx_ready_for_review"
        if not failed
        else "blocked_by_synthetic_rate_card_sandbox_xlsx"
    )
    report = {
        "schema_version": "0.1",
        "synthetic_rate_card_sandbox_xlsx_export_id": "synratesandboxxlsx-"
        + digest_json({"source": source_hash, "package": package_hash}).removeprefix("sha256:")[
            :16
        ],
        "status": status,
        "methodology_version": METHODOLOGY_VERSION,
        "source_rate_card_ref": RATE_CARD_REF,
        "source_rate_card_sha256": source_hash,
        "candidate_package_sha256": package_hash,
        "candidate_package_filename": package_file.name,
        "pinned_rate_total": package.get("pinnedRateTotal") if isinstance(package, dict) else None,
        "draft_rate_total": package.get("draftRateTotal") if isinstance(package, dict) else None,
        "delta": package.get("delta") if isinstance(package, dict) else None,
        "changed_cell_count": package.get("changedCellCount")
        if isinstance(package, dict)
        else None,
        "cell_count": len(package.get("cells", []))
        if isinstance(package, dict) and isinstance(package.get("cells"), list)
        else 0,
        "checks": checks,
        "failed_check_count": len(failed),
        "workbook_filename": SANDBOX_EXPORT_WORKBOOK_FILENAME,
        "data_origin": "synthetic",
        "candidate_only": True,
        "non_authoritative": True,
        "local_output_only": True,
        "source_mutation_performed": False,
        "rate_card_applied_to_budget": False,
        "external_writes_performed": False,
        "lake_write_performed": False,
        "sqlite_write_performed": False,
        "budget_submission_authorized": False,
        "matter_opening_authorized": False,
        "silent_learning_performed": False,
        "generated_at": generated_at or now_iso(),
    }
    run_dir = Path(out_dir)
    run_dir.mkdir(parents=True, exist_ok=True)
    write_json(run_dir / SANDBOX_EXPORT_REPORT_FILENAME, report)
    if not failed and isinstance(package, dict):
        _write_workbook(
            run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME,
            package=package,
            source_hash=source_hash,
            package_hash=package_hash,
            source_cells=source_cells,
        )
    return report, run_dir
