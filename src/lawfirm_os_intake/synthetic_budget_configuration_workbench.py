"""Inventory editable synthetic budget inputs without creating pricing authority."""

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import yaml

from .models import (
    SyntheticBudgetConfigurationEntry,
    SyntheticBudgetConfigurationSource,
    SyntheticBudgetConfigurationWorkbenchCheck,
    SyntheticBudgetConfigurationWorkbenchReport,
)
from .util import digest_json, digest_text, now_iso, write_json

REPORT_FILENAME = "synthetic_budget_configuration_workbench_report.json"
MARKDOWN_FILENAME = "synthetic_budget_configuration_workbench.md"
WORKBOOK_FILENAME = "synthetic_budget_configuration_workbench.xlsx"
METHODOLOGY_VERSION = "synthetic_budget_configuration_workbench.v0_1"
SOURCE_SPECS = (
    ("practice_profile", "context/synthetic-profiles/insurance-defense.yaml", "practice_profile"),
    ("rate_card", "config/synthetic-carrier-rate-card.yaml", "rate_card"),
    ("guideline", "config/synthetic-carrier-guideline.yaml", "guideline"),
    (
        "nonlinear_template",
        "examples/synthetic/labor-employment/labor-employment-nonlinear-budget-templates.json",
        "nonlinear_template",
    ),
)


def _check(check_id: str, passed: bool, message: str, *refs: str):
    return SyntheticBudgetConfigurationWorkbenchCheck(
        check_id=check_id,
        status="passed" if passed else "failed",
        message=message,
        evidence_refs=list(refs),
    )


def _mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number >= 0 else None


def _safe_text(value: str) -> str:
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def _truthy_real_data_flags(value: Any, prefix: str = "") -> list[str]:
    findings: list[str] = []
    if isinstance(value, dict):
        for key, nested in value.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            if str(key).startswith("contains_real_") and nested is True:
                findings.append(path)
            findings.extend(_truthy_real_data_flags(nested, path))
    elif isinstance(value, list):
        for index, nested in enumerate(value):
            findings.extend(_truthy_real_data_flags(nested, f"{prefix}[{index}]"))
    return findings


def _read_sources(root: Path) -> tuple[dict[str, tuple[str, dict[str, Any]]], list[Any]]:
    loaded: dict[str, tuple[str, dict[str, Any]]] = {}
    sources = []
    for source_id, source_ref, source_kind in SOURCE_SPECS:
        path = root / source_ref
        text = path.read_text(encoding="utf-8") if path.is_file() else ""
        parsed = json.loads(text) if source_ref.endswith(".json") and text else yaml.safe_load(text)
        payload = _mapping(parsed)
        loaded[source_id] = (text, payload)
        sources.append(
            SyntheticBudgetConfigurationSource(
                source_id=source_id,
                source_ref=source_ref,
                source_sha256=digest_text(text),
                source_kind=source_kind,
            )
        )
    return loaded, sources


def _entries(loaded: dict[str, tuple[str, dict[str, Any]]]) -> tuple[list[Any], list[str]]:
    entries: list[Any] = []
    invalid_paths: list[str] = []

    def add(
        source_id: str,
        source_ref: str,
        config_path: str,
        label: str,
        raw_value: Any,
        unit: str,
        math_effect: str,
    ) -> None:
        value = _numeric(raw_value)
        if value is None:
            invalid_paths.append(config_path)
            return
        entries.append(
            SyntheticBudgetConfigurationEntry(
                entry_id="synconfig-"
                + digest_json([source_id, config_path]).removeprefix("sha256:")[:16],
                source_id=source_id,
                source_ref=source_ref,
                config_path=config_path,
                label=label,
                value=value,
                unit=unit,
                math_effect=math_effect,
            )
        )

    profile_ref = SOURCE_SPECS[0][1]
    profile = loaded["practice_profile"][1]
    for role, value in sorted(_mapping(profile.get("synthetic_hourly_rates")).items()):
        add(
            "practice_profile",
            profile_ref,
            f"synthetic_hourly_rates.{role}",
            f"Fallback {role} rate",
            value,
            "hourly_rate",
            "proposal_rate_fallback",
        )
    for family, template in sorted(_mapping(profile.get("budget_templates")).items()):
        template_map = _mapping(template)
        add(
            "practice_profile",
            profile_ref,
            f"budget_templates.{family}.contingency_percent",
            f"{family} contingency",
            template_map.get("contingency_percent"),
            "percent",
            "proposal_contingency",
        )
        for phase in template_map.get("phases", []):
            phase_map = _mapping(phase)
            phase_id = str(phase_map.get("phase_id", "unknown"))
            for task in phase_map.get("tasks", []):
                task_map = _mapping(task)
                task_id = str(task_map.get("task_id", "unknown"))
                prefix = f"budget_templates.{family}.phases[{phase_id}].tasks[{task_id}]"
                label = f"{family} {phase_id}/{task_id}"
                add(
                    "practice_profile",
                    profile_ref,
                    f"{prefix}.estimated_hours",
                    f"{label} hours",
                    task_map.get("estimated_hours"),
                    "hours",
                    "proposal_template_hours",
                )
                add(
                    "practice_profile",
                    profile_ref,
                    f"{prefix}.estimated_expenses",
                    f"{label} expenses",
                    task_map.get("estimated_expenses"),
                    "currency",
                    "proposal_template_expense",
                )
                if "hours_per_unit" in task_map:
                    add(
                        "practice_profile",
                        profile_ref,
                        f"{prefix}.hours_per_unit",
                        f"{label} driver hours per unit",
                        task_map.get("hours_per_unit"),
                        "hours",
                        "proposal_template_hours",
                    )
                if "expense_per_unit" in task_map:
                    add(
                        "practice_profile",
                        profile_ref,
                        f"{prefix}.expense_per_unit",
                        f"{label} driver expense per unit",
                        task_map.get("expense_per_unit"),
                        "currency",
                        "proposal_template_expense",
                    )

    rate_ref = SOURCE_SPECS[1][1]
    card = loaded["rate_card"][1]
    for carrier_id, carrier in sorted(_mapping(card.get("carriers")).items()):
        carrier_map = _mapping(carrier)
        for state, roles in sorted(_mapping(carrier_map.get("schedule")).items()):
            for role, value in sorted(_mapping(roles).items()):
                add(
                    "rate_card",
                    rate_ref,
                    f"carriers.{carrier_id}.schedule.{state}.{role}",
                    f"{carrier_id} {state} {role} rate",
                    value,
                    "hourly_rate",
                    "proposal_rate_fallback",
                )
        for timekeeper_id, override in sorted(
            _mapping(carrier_map.get("named_timekeeper_overrides")).items()
        ):
            add(
                "rate_card",
                rate_ref,
                f"carriers.{carrier_id}.named_timekeeper_overrides.{timekeeper_id}.approved_rate",
                f"{carrier_id} named timekeeper override",
                _mapping(override).get("approved_rate"),
                "hourly_rate",
                "named_timekeeper_override",
            )

    guideline_ref = SOURCE_SPECS[2][1]
    guideline = loaded["guideline"][1]
    for carrier_id, carrier in sorted(_mapping(guideline.get("carriers")).items()):
        carrier_map = _mapping(carrier)
        add(
            "guideline",
            guideline_ref,
            f"carriers.{carrier_id}.variance_approval_percent",
            f"{carrier_id} variance approval threshold",
            carrier_map.get("variance_approval_percent"),
            "percent",
            "guideline_variance_threshold",
        )
        for role, value in sorted(_mapping(carrier_map.get("rate_caps")).items()):
            add(
                "guideline",
                guideline_ref,
                f"carriers.{carrier_id}.rate_caps.{role}",
                f"{carrier_id} {role} rate cap",
                value,
                "hourly_rate",
                "guideline_projection_rate_cap",
            )
        for code, value in sorted(_mapping(carrier_map.get("expense_caps")).items()):
            add(
                "guideline",
                guideline_ref,
                f"carriers.{carrier_id}.expense_caps.{code}",
                f"{carrier_id} {code} expense cap",
                value,
                "currency",
                "guideline_projection_expense_cap",
            )
        for threshold, value in sorted(
            _mapping(carrier_map.get("pre_approval_thresholds")).items()
        ):
            unit = (
                "currency"
                if threshold.endswith("amount")
                else "count"
                if threshold.endswith("count")
                else "hours"
            )
            add(
                "guideline",
                guideline_ref,
                f"carriers.{carrier_id}.pre_approval_thresholds.{threshold}",
                f"{carrier_id} {threshold}",
                value,
                unit,
                "guideline_preapproval_threshold",
            )
    return entries, invalid_paths


def build_synthetic_budget_configuration_workbench_report(
    *, repo_root: str | Path, generated_at: str | None = None
) -> SyntheticBudgetConfigurationWorkbenchReport:
    root = Path(repo_root)
    source_paths = [root / source_ref for _, source_ref, _ in SOURCE_SPECS]
    source_bytes_before = {path: path.read_bytes() for path in source_paths if path.is_file()}
    loaded, sources = _read_sources(root)
    entries, invalid_paths = _entries(loaded)
    source_files_present = all((root / source.source_ref).is_file() for source in sources)
    profile, card, guideline, nonlinear = (loaded[key][1] for key, _, _ in SOURCE_SPECS)
    real_flags = [
        *(_truthy_real_data_flags(profile, "practice_profile")),
        *(_truthy_real_data_flags(card, "rate_card")),
        *(_truthy_real_data_flags(guideline, "guideline")),
        *(_truthy_real_data_flags(nonlinear, "nonlinear_template")),
    ]
    synthetic_declarations = (
        profile.get("contains_real_firm_data") is False
        and card.get("data_origin") == "synthetic"
        and card.get("candidate_only") is True
        and card.get("real_rate_import_allowed") is False
        and guideline.get("data_scope") == "synthetic_only"
        and guideline.get("contains_real_carrier_guidelines") is False
        and nonlinear.get("data_origin") == "synthetic"
        and not real_flags
    )
    effect_counts = dict(sorted(Counter(entry.math_effect for entry in entries).items()))
    source_inputs_unchanged = all(
        path.is_file() and path.read_bytes() == content
        for path, content in source_bytes_before.items()
    )
    checks = [
        _check(
            "source_files_present",
            source_files_present,
            "Every declared editable synthetic source must be present.",
            *[source.source_ref for source in sources],
        ),
        _check(
            "synthetic_declarations_hold",
            synthetic_declarations,
            "All editable sources must remain synthetic candidate-only and block real-rate import.",
            *[source.source_ref for source in sources],
            *real_flags,
        ),
        _check(
            "numeric_entries_nonnegative",
            not invalid_paths,
            "Every listed numeric configuration input must be a nonnegative number.",
            *invalid_paths,
        ),
        _check(
            "configuration_paths_unique",
            len({entry.config_path for entry in entries}) == len(entries),
            "Each editable numeric field must have one unambiguous source path.",
        ),
        _check(
            "all_math_effects_explicit",
            bool(effect_counts) and all(entry.math_effect for entry in entries),
            "Every inventory value must state whether it affects proposal, projection, threshold, or override behavior.",
        ),
        _check(
            "nonlinear_template_is_structural_only",
            bool(nonlinear.get("templates"))
            and not any(entry.source_id == "nonlinear_template" for entry in entries),
            "The L&E nonlinear template is inventoried as a structural source, not fabricated as a numeric pricing input.",
            SOURCE_SPECS[3][1],
        ),
        _check(
            "source_inputs_unchanged_during_build",
            source_inputs_unchanged,
            "Every declared editable synthetic source must remain byte-identical while the inventory builds.",
            *[source.source_ref for source in sources],
        ),
    ]
    failed = sum(check.status == "failed" for check in checks)
    basis = {
        "sources": [(source.source_id, source.source_sha256) for source in sources],
        "entries": [entry.entry_id for entry in entries],
        "methodology_version": METHODOLOGY_VERSION,
    }
    return SyntheticBudgetConfigurationWorkbenchReport(
        synthetic_budget_configuration_workbench_report_id="synbudgetconfig-"
        + digest_json(basis).removeprefix("sha256:")[:16],
        status="synthetic_budget_configuration_workbench_ready_for_review"
        if not failed
        else "blocked_by_synthetic_budget_configuration_workbench",
        sources=sources,
        entries=entries,
        source_count=len(sources),
        entry_count=len(entries),
        entries_by_math_effect=effect_counts,
        checks=checks,
        failed_check_count=failed,
        workbook_filename=WORKBOOK_FILENAME,
        markdown_filename=MARKDOWN_FILENAME,
        display_banner={
            "summary": "Synthetic editable-input inventory only. Workbook edits are not imported; update the checked-in synthetic source, then regenerate.",
            "candidate_only": True,
            "synthetic_only": True,
            "read_only_ui": True,
            "blocked_actions": [
                "real_rate_import",
                "workbook_import",
                "browser_side_pricing",
                "budget_submission",
                "exception_lake_write",
                "sqlite_write",
                "silent_learning",
            ],
        },
        required_next_gates=[
            "human_review_before_synthetic_configuration_change",
            "legal_knowledge_runtime_reviewed_snapshot_before_real_rate_replacement",
            "orchestrator_owned_edit_session_before_any_runtime_import",
        ],
        generated_at=generated_at or now_iso(),
    )


def _write_workbook(report: SyntheticBudgetConfigurationWorkbenchReport, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    read_me = workbook.active
    read_me.title = "Read Me"
    for row in [
        ["Synthetic Budget Configuration Worksheet"],
        ["Status", report.status],
        ["Boundary", report.display_banner["summary"]],
        [
            "How to apply",
            "Copy reviewed synthetic values into the declared YAML/JSON source, then run this builder again.",
        ],
        ["Import behavior", "None. This workbook is never read by runtime budget logic."],
        ["Workbook formulas", "none"],
    ]:
        read_me.append(row)
    read_me.column_dimensions["A"].width = 22
    read_me.column_dimensions["B"].width = 118
    values = workbook.create_sheet("Editable Values")
    values.append(
        [
            "Synthetic candidate-only worksheet. Yellow cells are copy/reference values, not an import surface."
        ]
    )
    values.append([])
    values.append(["Source", "Config Path", "Label", "Value", "Unit", "Math Effect"])
    for entry in report.entries:
        values.append(
            [
                _safe_text(entry.source_ref),
                _safe_text(entry.config_path),
                _safe_text(entry.label),
                entry.value,
                entry.unit,
                entry.math_effect,
            ]
        )
    header_fill = PatternFill("solid", fgColor="234157")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in values[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for cell in values["D"][3:]:
        cell.fill = input_fill
    values.freeze_panes = "A4"
    values.auto_filter.ref = (
        f"A3:{values.cell(row=values.max_row, column=values.max_column).coordinate}"
    )
    for column, width in {"A": 62, "B": 82, "C": 54, "D": 18, "E": 18, "F": 36}.items():
        values.column_dimensions[column].width = width
    manifest = workbook.create_sheet("Source Manifest")
    manifest.append(["Source", "Path", "SHA-256", "Kind", "Editable"])
    for source in report.sources:
        manifest.append(
            [source.source_id, source.source_ref, source.source_sha256, source.source_kind, "true"]
        )
    for cell in manifest[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for column, width in {"A": 24, "B": 94, "C": 76, "D": 24, "E": 12}.items():
        manifest.column_dimensions[column].width = width
    workbook.properties.title = "Synthetic Budget Configuration Worksheet"
    workbook.properties.subject = "Synthetic candidate-only local evidence"
    workbook.properties.creator = "LawFirm-os-intake"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    with ZipFile(path) as archive:
        if any("vbaProject" in name or "externalLinks/" in name for name in archive.namelist()):
            raise ValueError("active workbook content prohibited")


def run_synthetic_budget_configuration_workbench(
    *, repo_root: str | Path, out_dir: str | Path, generated_at: str | None = None
):
    report = build_synthetic_budget_configuration_workbench_report(
        repo_root=repo_root, generated_at=generated_at
    )
    target = Path(out_dir)
    target.mkdir(parents=True, exist_ok=True)
    write_json(target / REPORT_FILENAME, report.model_dump(mode="json"))
    (target / MARKDOWN_FILENAME).write_text(
        "# Synthetic Budget Configuration Workbench\n\n" + report.display_banner["summary"] + "\n",
        encoding="utf-8",
    )
    if report.status.endswith("ready_for_review"):
        _write_workbook(report, target / WORKBOOK_FILENAME)
    return report, target
