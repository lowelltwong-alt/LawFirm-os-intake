"""Render a budget proposal into a UTBMS litigation budget form workbook.

Two modes:

- synthetic (``template_path=None``): build a clean UTBMS budget workbook that mirrors the
  carrier form layout (Phase/Task rows, an Original Budgeted Amount column, phase
  subtotals, and a grand total);
- fill-existing (``template_path`` given): open a copy of an existing UTBMS budget form,
  locate each UTBMS code in the Phase/Task column, and write the amount into the
  Original Budgeted Amount column - preserving the workbook's own subtotal/total formulas.

The form is "Original Budgeted Amount" only, matching the attorney's fill-out duty. The
carrier form has no contingency line, so the form total is fees + expenses; the model's
contingency stays in the JSON proposal, not the form. The output is a proposal and is not
authorized for submission. ``openpyxl`` is imported lazily so importing this module does
not require it.
"""

from __future__ import annotations

from hashlib import sha256
import re
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from .models import BudgetProposal
from .models import (
    BudgetFormCodeMapping,
    BudgetFormFormulaCheck,
    BudgetFormMappingReport,
    BudgetFormTemplateAuditReport,
)
from .util import new_id, now_iso

AMOUNT_HEADER = "Original Budgeted Amount"
TASK_HEADER = "Phase / Task"
AMOUNT_BILLED_HEADER = "Amount Billed to Date"
AMOUNT_REMAINING_HEADER = "Original Budget Amount Remaining"
TOTAL_LABEL = "Total Budgeted ($)"
_CODE_RE = re.compile(r"\((L\d{3}|E\d{3})\)")
_CELL_REF_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")

# The UTBMS litigation phase/task and expense code skeleton, in form order.
UTBMS_FORM_ROWS: list[tuple[str, str, str]] = [
    ("L100", "Case Assessment, Development and Administration", "phase"),
    ("L110", "Fact Investigation / Development", "task"),
    ("L120", "Analysis / Strategy", "task"),
    ("L130", "Experts / Consultants", "task"),
    ("L140", "Document / File Management", "task"),
    ("L150", "Budgeting", "task"),
    ("L160", "Settlement / Non-Binding ADR", "task"),
    ("L190", "Other Case Assessment, Development and Administration", "task"),
    ("L200", "Pre-Trial Pleading and Motions", "phase"),
    ("L210", "Pleading", "task"),
    ("L220", "Preliminary Injunctions / Provisional Remedies", "task"),
    ("L230", "Court Mandated Conferences", "task"),
    ("L240", "Dispositive Motions", "task"),
    ("L250", "Other Written Motions and Sanctions", "task"),
    ("L260", "Class Action Certification and Notice", "task"),
    ("L300", "Discovery", "phase"),
    ("L310", "Written Discovery", "task"),
    ("L320", "Document Production", "task"),
    ("L330", "Depositions", "task"),
    ("L340", "Expert Discovery", "task"),
    ("L350", "Discovery Motions", "task"),
    ("L390", "Other Discovery", "task"),
    ("L400", "Trial Preparation and Trial", "phase"),
    ("L410", "Fact Witnesses", "task"),
    ("L420", "Expert Witnesses", "task"),
    ("L430", "Written Motions and Submissions", "task"),
    ("L440", "Other Trial Preparation and Support", "task"),
    ("L450", "Trial and Hearing Attendance", "task"),
    ("L460", "Post-Trial Motions and Submissions", "task"),
    ("L470", "Enforcement", "task"),
    ("L500", "Appeal", "phase"),
    ("L510", "Appellate Motions and Submissions", "task"),
    ("L520", "Appellate Briefs", "task"),
    ("L530", "Oral Argument", "task"),
    ("E100", "Expenses", "phase"),
    ("E101", "Copying", "task"),
    ("E102", "Outside Printing", "task"),
    ("E103", "Word Processing", "task"),
    ("E104", "Facsimile", "task"),
    ("E105", "Telephone", "task"),
    ("E106", "Online Research", "task"),
    ("E107", "Messengers / Overnite", "task"),
    ("E108", "Postage", "task"),
    ("E109", "Local Travel", "task"),
    ("E110", "Out-of-Town Travel", "task"),
    ("E111", "Meals", "task"),
    ("E112", "Court Fees", "task"),
    ("E113", "Subpoena Fees", "task"),
    ("E114", "Witness Fees", "task"),
    ("E115", "Court Reporting & Transcripts", "task"),
    ("E116", "Trial Transcripts", "task"),
    ("E117", "Trial Exhibits", "task"),
    ("E118", "Litigation, Support Vendors", "task"),
    ("E119", "Experts", "task"),
    ("E120", "Private Investigators", "task"),
    ("E121", "Arbitrators / Mediators", "task"),
    ("E122", "Local Counsel", "task"),
    ("E123", "Other Professionals", "task"),
    ("E124", "Other", "task"),
]

# Each phase code mapped to its task codes, for subtotal computation.
_PHASE_TASKS: dict[str, list[str]] = {}
_current_phase = None
for _code, _label, _kind in UTBMS_FORM_ROWS:
    if _kind == "phase":
        _current_phase = _code
        _PHASE_TASKS[_code] = []
    elif _current_phase is not None:
        _PHASE_TASKS[_current_phase].append(_code)

_ROW_KINDS = {code: kind for code, _, kind in UTBMS_FORM_ROWS}
_REQUIRED_CODES = [code for code, _, _ in UTBMS_FORM_ROWS]
_TEMPLATE_CHECKLIST_ITEMS = [
    "Use the active worksheet that contains the carrier UTBMS budget form.",
    "Keep the Phase / Task column and Original Budgeted Amount column visible and labeled.",
    "Keep every expected UTBMS L/E phase and task code present exactly once.",
    "Keep Total Budgeted ($) tied to original-budget phase subtotal cells.",
    "Keep original-budget phase cells tied to their original-budget task cells.",
    "Keep task remaining formulas tied to original budget minus billed-to-date where present.",
    "Do not add real client data, matter data, private rates, or carrier submission state.",
]


def form_amounts(budget: BudgetProposal) -> dict[str, float]:
    """Amount per UTBMS code: L-codes carry fees, E-codes carry mapped expenses."""

    amounts: dict[str, float] = {}
    for line in budget.lines:
        code = line.external_code_candidate
        if code and code.startswith("L"):
            amounts[code] = round(amounts.get(code, 0.0) + (line.estimated_fees or 0.0), 2)
        expense_code = getattr(line, "expense_code", None)
        if expense_code and line.estimated_expenses:
            amounts[expense_code] = round(
                amounts.get(expense_code, 0.0) + float(line.estimated_expenses), 2
            )
    return amounts


def _phase_subtotals(amounts: dict[str, float]) -> dict[str, float]:
    return {
        phase: round(sum(amounts.get(task, 0.0) for task in tasks), 2)
        for phase, tasks in _PHASE_TASKS.items()
    }


def _file_sha256(path: Path) -> str:
    return "sha256:" + sha256(path.read_bytes()).hexdigest()


def _check(
    check_id: str,
    passed: bool,
    message: str,
    *,
    cell: str | None = None,
    actual_formula: str | None = None,
    expected_refs: list[str] | None = None,
    actual_refs: list[str] | None = None,
    warning: bool = False,
) -> BudgetFormFormulaCheck:
    return BudgetFormFormulaCheck(
        check_id=check_id,
        status="warning" if warning else ("passed" if passed else "failed"),
        message=message,
        cell=cell,
        actual_formula=actual_formula,
        expected_refs=expected_refs or [],
        actual_refs=actual_refs or [],
    )


def _cell_refs(formula: Any) -> list[str]:
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    refs = {f"{col.upper()}{row}" for col, row in _CELL_REF_RE.findall(formula)}
    return sorted(refs)


def _formula_matches_refs(formula: Any, expected_refs: list[str]) -> tuple[bool, list[str]]:
    actual_refs = _cell_refs(formula)
    return sorted(expected_refs) == actual_refs, actual_refs


def _find_headers(ws: Any) -> dict[str, tuple[int, int]]:
    headers: dict[str, tuple[int, int]] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            normalized = value.strip()
            if normalized == AMOUNT_HEADER:
                headers[AMOUNT_HEADER] = (cell.row, cell.column)
            elif normalized == AMOUNT_BILLED_HEADER:
                headers[AMOUNT_BILLED_HEADER] = (cell.row, cell.column)
            elif normalized == AMOUNT_REMAINING_HEADER:
                headers[AMOUNT_REMAINING_HEADER] = (cell.row, cell.column)
            elif TASK_HEADER in normalized:
                headers[TASK_HEADER] = (cell.row, cell.column)
    return headers


def _find_total_cell(ws: Any) -> str | None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == TOTAL_LABEL:
                return ws.cell(row=cell.row, column=cell.column + 1).coordinate
    return None


def _template_code_rows(ws: Any, task_col: int | None) -> dict[str, list[int]]:
    rows: dict[str, list[int]] = {}
    if task_col is None:
        return rows
    for row_index in range(1, ws.max_row + 1):
        value = ws.cell(row=row_index, column=task_col).value
        if not isinstance(value, str):
            continue
        match = _CODE_RE.search(value)
        if match is None:
            continue
        rows.setdefault(match.group(1), []).append(row_index)
    return rows


def _build_code_mappings(
    ws: Any,
    code_rows: dict[str, list[int]],
    amount_col: int | None,
    task_col: int | None,
    amounts: dict[str, float],
) -> list[BudgetFormCodeMapping]:
    mappings: list[BudgetFormCodeMapping] = []
    if amount_col is None or task_col is None:
        return mappings
    for code in _REQUIRED_CODES:
        rows = code_rows.get(code, [])
        if len(rows) != 1:
            continue
        row_index = rows[0]
        mappings.append(
            BudgetFormCodeMapping(
                code=code,
                kind=_ROW_KINDS[code],  # type: ignore[arg-type]
                row=row_index,
                label=str(ws.cell(row=row_index, column=task_col).value or ""),
                amount_cell=ws.cell(row=row_index, column=amount_col).coordinate,
                amount=round(amounts.get(code, 0.0), 2),
            )
        )
    return mappings


def _formula_checks(
    ws: Any,
    code_rows: dict[str, list[int]],
    *,
    amount_col: int | None,
    billed_col: int | None,
    remaining_col: int | None,
    total_cell: str | None,
) -> list[BudgetFormFormulaCheck]:
    checks: list[BudgetFormFormulaCheck] = []
    if amount_col is None:
        return [
            _check(
                "original_budget_amount_column_present",
                False,
                "Template must contain the Original Budgeted Amount column.",
            )
        ]

    if total_cell is None:
        checks.append(
            _check(
                "original_budget_total_formula",
                False,
                "Template must contain a Total Budgeted ($) cell next to the total label.",
            )
        )
    else:
        expected_refs = [
            ws.cell(row=code_rows[phase][0], column=amount_col).coordinate
            for phase, tasks in _PHASE_TASKS.items()
            if tasks and len(code_rows.get(phase, [])) == 1
        ]
        formula = ws[total_cell].value
        passed, actual_refs = _formula_matches_refs(formula, expected_refs)
        checks.append(
            _check(
                "original_budget_total_formula",
                passed,
                "Original budget total must sum original-budget phase cells.",
                cell=total_cell,
                actual_formula=formula if isinstance(formula, str) else None,
                expected_refs=expected_refs,
                actual_refs=actual_refs,
            )
        )

    for phase, tasks in _PHASE_TASKS.items():
        if len(code_rows.get(phase, [])) != 1:
            continue
        phase_cell = ws.cell(row=code_rows[phase][0], column=amount_col)
        expected_refs = [
            ws.cell(row=code_rows[task][0], column=amount_col).coordinate
            for task in tasks
            if len(code_rows.get(task, [])) == 1
        ]
        formula = phase_cell.value
        passed, actual_refs = _formula_matches_refs(formula, expected_refs)
        checks.append(
            _check(
                f"phase_{phase.lower()}_original_budget_formula",
                passed,
                f"Phase {phase} original-budget cell must sum its task cells.",
                cell=phase_cell.coordinate,
                actual_formula=formula if isinstance(formula, str) else None,
                expected_refs=expected_refs,
                actual_refs=actual_refs,
            )
        )

    if billed_col is None or remaining_col is None:
        checks.append(
            _check(
                "task_remaining_formula_columns_present",
                False,
                "Template must contain billed-to-date and original-budget remaining columns.",
            )
        )
        return checks

    for code, rows in code_rows.items():
        if _ROW_KINDS.get(code) != "task" or len(rows) != 1:
            continue
        row_index = rows[0]
        remaining_cell = ws.cell(row=row_index, column=remaining_col)
        formula = remaining_cell.value
        if formula in (None, ""):
            continue
        amount_ref = ws.cell(row=row_index, column=amount_col).coordinate
        billed_ref = ws.cell(row=row_index, column=billed_col).coordinate
        expected_refs = [amount_ref, billed_ref]
        passed, actual_refs = _formula_matches_refs(formula, expected_refs)
        checks.append(
            _check(
                f"task_{code.lower()}_remaining_formula",
                passed,
                f"Task {code} remaining cell must reference original budget and billed-to-date.",
                cell=remaining_cell.coordinate,
                actual_formula=formula if isinstance(formula, str) else None,
                expected_refs=expected_refs,
                actual_refs=actual_refs,
            )
        )
    return checks


def build_budget_form_mapping_report(
    budget: BudgetProposal,
    template_path: str | Path,
) -> BudgetFormMappingReport:
    from openpyxl import load_workbook

    template = Path(template_path)
    wb = load_workbook(template, data_only=False)
    ws = wb.active
    headers = _find_headers(ws)
    task_header = headers.get(TASK_HEADER)
    amount_header = headers.get(AMOUNT_HEADER)
    billed_header = headers.get(AMOUNT_BILLED_HEADER)
    remaining_header = headers.get(AMOUNT_REMAINING_HEADER)
    task_col = task_header[1] if task_header else None
    amount_col = amount_header[1] if amount_header else None
    billed_col = billed_header[1] if billed_header else None
    remaining_col = remaining_header[1] if remaining_header else None
    code_rows = _template_code_rows(ws, task_col)
    amounts = form_amounts(budget)
    duplicate_codes = sorted(code for code, rows in code_rows.items() if len(rows) > 1)
    missing_template_codes = sorted(code for code in _REQUIRED_CODES if code not in code_rows)
    missing_budget_mappings = sorted(code for code in amounts if code not in code_rows)
    unmapped_budget_amount_codes = sorted(set(missing_budget_mappings) | set(duplicate_codes))
    total_cell = _find_total_cell(ws)
    checks = _formula_checks(
        ws,
        code_rows,
        amount_col=amount_col,
        billed_col=billed_col,
        remaining_col=remaining_col,
        total_cell=total_cell,
    )
    structural_failures = [
        not task_header,
        not amount_header,
        bool(missing_template_codes),
        bool(duplicate_codes),
        bool(missing_budget_mappings),
        budget.not_authorized_for_client_submission is not True,
    ]
    status = (
        "passed"
        if not any(structural_failures) and all(check.status != "failed" for check in checks)
        else "failed"
    )
    return BudgetFormMappingReport(
        budget_form_mapping_report_id=new_id("budgetformmap"),
        budget_proposal_id=budget.budget_proposal_id,
        status=status,
        template_sha256=_file_sha256(template),
        sheet_name=ws.title,
        task_header_cell=(
            ws.cell(row=task_header[0], column=task_header[1]).coordinate if task_header else None
        ),
        amount_header_cell=(
            ws.cell(row=amount_header[0], column=amount_header[1]).coordinate
            if amount_header
            else None
        ),
        total_cell=total_cell,
        task_column=task_col,
        amount_column=amount_col,
        code_mappings=_build_code_mappings(ws, code_rows, amount_col, task_col, amounts),
        amounts_by_code=dict(sorted(amounts.items())),
        l_code_total=round(sum(v for k, v in amounts.items() if k.startswith("L")), 2),
        e_code_total=round(sum(v for k, v in amounts.items() if k.startswith("E")), 2),
        missing_template_codes=missing_template_codes,
        duplicate_template_codes=duplicate_codes,
        missing_budget_mappings=missing_budget_mappings,
        unmapped_budget_amount_codes=unmapped_budget_amount_codes,
        formula_checks=checks,
        warnings=[],
        not_authorized_for_client_submission=budget.not_authorized_for_client_submission,
        generated_at=now_iso(),
    )


def build_budget_form_template_audit_report(
    template_path: str | Path,
) -> BudgetFormTemplateAuditReport:
    """Audit a carrier-style UTBMS budget form template without a budget proposal."""

    from openpyxl import load_workbook

    template = Path(template_path)
    wb = load_workbook(template, data_only=False)
    ws = wb.active
    headers = _find_headers(ws)
    task_header = headers.get(TASK_HEADER)
    amount_header = headers.get(AMOUNT_HEADER)
    billed_header = headers.get(AMOUNT_BILLED_HEADER)
    remaining_header = headers.get(AMOUNT_REMAINING_HEADER)
    task_col = task_header[1] if task_header else None
    amount_col = amount_header[1] if amount_header else None
    billed_col = billed_header[1] if billed_header else None
    remaining_col = remaining_header[1] if remaining_header else None
    code_rows = _template_code_rows(ws, task_col)
    duplicate_codes = sorted(code for code, rows in code_rows.items() if len(rows) > 1)
    missing_template_codes = sorted(code for code in _REQUIRED_CODES if code not in code_rows)
    total_cell = _find_total_cell(ws)
    checks = _formula_checks(
        ws,
        code_rows,
        amount_col=amount_col,
        billed_col=billed_col,
        remaining_col=remaining_col,
        total_cell=total_cell,
    )
    structural_failures = [
        not task_header,
        not amount_header,
        bool(missing_template_codes),
        bool(duplicate_codes),
    ]
    status = (
        "passed"
        if not any(structural_failures) and all(check.status != "failed" for check in checks)
        else "failed"
    )
    return BudgetFormTemplateAuditReport(
        budget_form_template_audit_report_id=new_id("budgetformtemplateaudit"),
        status=status,
        template_sha256=_file_sha256(template),
        sheet_name=ws.title,
        task_header_cell=(
            ws.cell(row=task_header[0], column=task_header[1]).coordinate if task_header else None
        ),
        amount_header_cell=(
            ws.cell(row=amount_header[0], column=amount_header[1]).coordinate
            if amount_header
            else None
        ),
        total_cell=total_cell,
        task_column=task_col,
        amount_column=amount_col,
        code_mappings=_build_code_mappings(ws, code_rows, amount_col, task_col, {}),
        missing_template_codes=missing_template_codes,
        duplicate_template_codes=duplicate_codes,
        formula_checks=checks,
        checklist_items=_TEMPLATE_CHECKLIST_ITEMS,
        warnings=[],
        generated_at=now_iso(),
    )


def enforce_budget_form_mapping_report(report: BudgetFormMappingReport) -> None:
    if report.status == "passed":
        return
    failed_checks = [check.check_id for check in report.formula_checks if check.status == "failed"]
    reasons = [
        *failed_checks,
        *[f"missing_template_code:{code}" for code in report.missing_template_codes],
        *[f"duplicate_template_code:{code}" for code in report.duplicate_template_codes],
        *[f"missing_budget_mapping:{code}" for code in report.missing_budget_mappings],
    ]
    if report.not_authorized_for_client_submission is not True:
        reasons.append("budget_authorized_for_submission")
    raise ValueError("budget form mapping failed: " + ", ".join(reasons))


def enforce_budget_form_template_audit_report(
    report: BudgetFormTemplateAuditReport,
) -> None:
    if report.status == "passed":
        return
    failed_checks = [check.check_id for check in report.formula_checks if check.status == "failed"]
    reasons = [
        *failed_checks,
        *[f"missing_template_code:{code}" for code in report.missing_template_codes],
        *[f"duplicate_template_code:{code}" for code in report.duplicate_template_codes],
    ]
    raise ValueError("budget form template audit failed: " + ", ".join(reasons))


def _build_synthetic_form(budget: BudgetProposal, amounts: dict[str, float], out_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    subtotals = _phase_subtotals(amounts)
    grand_total = round(sum(subtotals.values()), 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Form"
    bold = Font(bold=True)
    ws["A1"] = "BUDGET FORM"
    ws["A1"].font = bold
    ws["A2"] = (
        "Synthetic budget proposal - Original Budgeted Amount only. Proposal for human "
        "review; not authorized for client or carrier submission."
    )
    ws["A3"] = f"Budget proposal id: {budget.budget_proposal_id}"
    ws["A4"] = f"Matter family: {budget.matter_family} | currency: {budget.currency}"

    header_row = 6
    headers = [
        TASK_HEADER,
        AMOUNT_HEADER,
        "Amount Billed to Date",
        "Original Budget Amount Remaining",
        "New Budgeted Amount",
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = bold

    row = header_row + 1
    for code, label, kind in UTBMS_FORM_ROWS:
        if kind == "phase":
            ws.cell(row=row, column=1, value=f"({code}) {label}").font = bold
            ws.cell(row=row, column=2, value=subtotals.get(code, 0.0)).font = bold
        else:
            ws.cell(row=row, column=1, value=f"     ({code}) {label}")
            amount = amounts.get(code)
            if amount:
                ws.cell(row=row, column=2, value=amount)
        row += 1

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="Total Budgeted ($)").font = bold
    ws.cell(row=total_row, column=2, value=grand_total).font = bold
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 22

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return grand_total


def _fill_existing_form(template_path: Path, amounts: dict[str, float], out_path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = wb.active
    amount_col: int | None = None
    task_col = 1
    scan_rows = min(ws.max_row, 40)
    for r in range(1, scan_rows + 1):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(row=r, column=c).value
            if not isinstance(value, str):
                continue
            if value.strip() == AMOUNT_HEADER:
                amount_col = c
            elif TASK_HEADER in value:
                task_col = c
        if amount_col is not None:
            break
    if amount_col is None:
        raise ValueError(f"template has no '{AMOUNT_HEADER}' column to fill")

    filled = 0
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=task_col).value
        if not isinstance(label, str):
            continue
        match = _CODE_RE.search(label)
        if match is None:
            continue
        code = match.group(1)
        if code in amounts:
            ws.cell(row=r, column=amount_col, value=amounts[code])
            filled += 1
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return filled


def render_budget_form(
    budget: BudgetProposal,
    out_path: str | Path,
    template_path: str | Path | None = None,
    mapping_report_out: str | Path | None = None,
) -> Path:
    """Render ``budget`` into a UTBMS budget form workbook at ``out_path``.

    With ``template_path`` the amounts are written into a copy of that existing form;
    without it, a synthetic UTBMS workbook is generated. Only fee/expense amounts are
    written; conflict, engagement, and submission remain out of scope.
    """

    out_path = Path(out_path)
    amounts = form_amounts(budget)
    if template_path is not None:
        report = build_budget_form_mapping_report(budget, template_path)
        if mapping_report_out is not None:
            Path(mapping_report_out).parent.mkdir(parents=True, exist_ok=True)
            Path(mapping_report_out).write_text(report.model_dump_json(indent=2), encoding="utf-8")
        enforce_budget_form_mapping_report(report)
        _fill_existing_form(Path(template_path), amounts, out_path)
    else:
        if mapping_report_out is not None:
            raise ValueError("--mapping-report-out requires --template")
        _build_synthetic_form(budget, amounts, out_path)
    return out_path
