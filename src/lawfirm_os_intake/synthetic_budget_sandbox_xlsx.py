"""Render validated browser sandbox drafts as local synthetic XLSX artifacts.

The browser export is an untrusted candidate package. This module validates it
against the pinned synthetic budget proposal before producing a workbook. It
never writes the source proposal or any runtime authority.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import ValidationError

from .models import BudgetProposal
from .synthetic_budget_input_workbench import BUDGET_PROPOSAL_REF
from .util import digest_json, digest_text, load_json, now_iso, write_json


SANDBOX_EXPORT_REPORT_FILENAME = "synthetic_budget_sandbox_xlsx_export_report.json"
SANDBOX_EXPORT_WORKBOOK_FILENAME = "synthetic_budget_sandbox_candidate.xlsx"
METHODOLOGY_VERSION = "synthetic_budget_sandbox_xlsx_export.v0_1"
REQUIRED_BLOCKED_ACTIONS = {
    "configuration_write",
    "real_rate_import",
    "budget_submission",
    "exception_lake_write",
    "sqlite_write",
    "silent_learning",
}
CENT = Decimal("0.01")


def _decimal(value: Any, *, cents: bool = False) -> Decimal:
    """Parse an untrusted numeric value without float or banker-rounding drift."""
    if isinstance(value, bool):
        raise ValueError("Boolean values are not valid budget amounts.")
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Budget values must be finite decimal numbers.") from exc
    if not decimal_value.is_finite():
        raise ValueError("Budget values must be finite decimal numbers.")
    if cents:
        rounded = decimal_value.quantize(CENT, rounding=ROUND_HALF_UP)
        if decimal_value != rounded:
            raise ValueError("Currency amounts must already have cent precision.")
        return rounded
    return decimal_value


def _money(value: Any) -> float:
    return float(_decimal(value, cents=True))


def _check(check_id: str, passed: bool, message: str, *refs: str) -> dict[str, Any]:
    return {
        "check_id": check_id,
        "status": "passed" if passed else "failed",
        "message": message,
        "evidence_refs": [ref for ref in refs if ref],
    }


def _source_lines(proposal: dict[str, Any]) -> list[dict[str, Any]]:
    lines = proposal.get("lines")
    return lines if isinstance(lines, list) else []


def _validate_package(
    package: Any, proposal: dict[str, Any], proposal_sha256: str
) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    if not isinstance(package, dict):
        return [_check("package_is_mapping", False, "Candidate package must be a JSON object.")]

    try:
        source_proposal = BudgetProposal.model_validate(proposal)
        source_proposal_schema_valid = True
        source_rates_synthetic = all(line.rate_is_synthetic for line in source_proposal.lines)
    except ValidationError:
        source_proposal_schema_valid = False
        source_rates_synthetic = False

    blocked_actions = package.get("blocked_actions")
    blocked_actions_valid = isinstance(blocked_actions, list) and all(
        isinstance(action, str) for action in blocked_actions
    )

    checks.extend(
        [
            _check(
                "source_proposal_schema_valid",
                source_proposal_schema_valid,
                "The pinned source proposal must satisfy the local budget proposal contract.",
                BUDGET_PROPOSAL_REF,
            ),
            _check(
                "source_rates_synthetic",
                source_rates_synthetic,
                "Every pinned source line must declare a synthetic rate before candidate rendering.",
                BUDGET_PROPOSAL_REF,
            ),
            _check(
                "candidate_synthetic_declaration",
                package.get("artifact_type") == "synthetic_budget_sandbox_change_package"
                and package.get("data_origin") == "synthetic"
                and package.get("candidate_only") is True
                and package.get("local_browser_draft") is True,
                "Package must declare the synthetic browser-draft candidate contract.",
            ),
            _check(
                "source_proposal_hash_matches",
                package.get("source_budget_proposal_sha256") == proposal_sha256,
                "Package must bind to the currently pinned synthetic proposal hash.",
                BUDGET_PROPOSAL_REF,
            ),
            _check(
                "blocked_actions_complete",
                blocked_actions_valid and REQUIRED_BLOCKED_ACTIONS <= set(blocked_actions),
                "Package must explicitly retain every no-write and no-submission boundary.",
            ),
        ]
    )

    source_lines = _source_lines(proposal)
    package_lines = package.get("lines")
    checks.append(
        _check(
            "line_count_matches_source",
            isinstance(package_lines, list) and len(package_lines) == len(source_lines),
            "Candidate package must contain every pinned proposal line exactly once.",
            BUDGET_PROPOSAL_REF,
        )
    )
    if not isinstance(package_lines, list):
        return checks

    seen: set[int] = set()
    line_math_ok = True
    nonnegative_lines_ok = True
    line_identity_ok = len(package_lines) == len(source_lines)
    for index, package_line in enumerate(package_lines):
        if not isinstance(package_line, dict):
            line_math_ok = False
            line_identity_ok = False
            continue
        line_number = package_line.get("lineNumber")
        if not isinstance(line_number, int) or line_number in seen:
            line_identity_ok = False
        else:
            seen.add(line_number)
            if index >= len(source_lines) or line_number != index + 1:
                line_identity_ok = False
        try:
            hours = _decimal(package_line["estimatedHours"])
            rate = _decimal(package_line["hourlyRate"], cents=True)
            expenses = _decimal(package_line["estimatedExpenses"], cents=True)
            fees = _decimal(package_line["estimatedFees"], cents=True)
            line_total = _decimal(package_line["lineTotal"], cents=True)
            if hours < 0 or rate < 0 or expenses < 0:
                nonnegative_lines_ok = False
            expected_fees = (hours * rate).quantize(CENT, rounding=ROUND_HALF_UP)
            if fees != expected_fees or line_total != fees + expenses:
                line_math_ok = False
        except (KeyError, TypeError, ValueError, InvalidOperation):
            line_math_ok = False
            nonnegative_lines_ok = False
    checks.extend(
        [
            _check(
                "line_identity_complete",
                line_identity_ok and len(seen) == len(source_lines),
                "Candidate line numbers must be unique and preserve the pinned proposal order.",
                BUDGET_PROPOSAL_REF,
            ),
            _check(
                "line_math_reconciles",
                line_math_ok,
                "Every candidate line must satisfy hours times rate plus expenses with cent precision.",
            ),
            _check(
                "nonnegative_candidate_amounts",
                nonnegative_lines_ok,
                "Candidate hours, rates, and expenses must be finite nonnegative values.",
            ),
        ]
    )

    try:
        source_total = _decimal(proposal["total_proposed_budget"], cents=True)
        package_pinned_total = _decimal(package["pinned_total"], cents=True)
        contingency = _decimal(package["fixed_contingency_amount"], cents=True)
        candidate_total = _decimal(package["draft_total"], cents=True)
        delta = _decimal(package["delta"], cents=True)
        line_total = sum(
            (_decimal(line["lineTotal"], cents=True) for line in package_lines),
            start=Decimal("0"),
        )
        nonnegative_totals_ok = all(
            amount >= 0 for amount in (package_pinned_total, contingency, candidate_total)
        )
        checks.extend(
            [
                _check(
                    "pinned_total_matches_source",
                    package_pinned_total == source_total,
                    "Package pinned total must equal the current synthetic proposal total.",
                    BUDGET_PROPOSAL_REF,
                ),
                _check(
                    "candidate_total_reconciles",
                    candidate_total == line_total + contingency,
                    "Candidate total must equal candidate line totals plus contingency.",
                ),
                _check(
                    "candidate_delta_reconciles",
                    delta == candidate_total - package_pinned_total,
                    "Candidate delta must equal candidate total minus pinned total.",
                ),
                _check(
                    "nonnegative_candidate_totals",
                    nonnegative_totals_ok,
                    "Pinned total, contingency, and candidate total must be nonnegative.",
                ),
            ]
        )
    except (KeyError, TypeError, ValueError, InvalidOperation):
        checks.extend(
            [
                _check(
                    "pinned_total_matches_source", False, "Package totals are missing or invalid."
                ),
                _check(
                    "candidate_total_reconciles", False, "Package totals are missing or invalid."
                ),
                _check(
                    "candidate_delta_reconciles", False, "Package totals are missing or invalid."
                ),
                _check(
                    "nonnegative_candidate_totals",
                    False,
                    "Package totals must be finite nonnegative values.",
                ),
            ]
        )
    return checks


def _write_workbook(
    path: Path,
    *,
    package: dict[str, Any],
    proposal: dict[str, Any],
    proposal_sha256: str,
    package_sha256: str,
) -> None:
    workbook = Workbook()
    read_me = workbook.active
    read_me.title = "Read Me"
    read_me.append(["Synthetic Budget Sandbox Candidate Workbook"])
    read_me.append([])
    for row in [
        ["Status", "candidate_only_local_export"],
        ["Source proposal", BUDGET_PROPOSAL_REF],
        ["Source proposal hash", proposal_sha256],
        ["Candidate package hash", package_sha256],
        ["Pinned total", package["pinned_total"]],
        ["Candidate total", package["draft_total"]],
        ["Candidate delta", package["delta"]],
        [
            "Boundary",
            "No source configuration write, real-rate import, submission, Lake/SQLite write, or silent learning.",
        ],
    ]:
        read_me.append(row)
    read_me["A1"].font = Font(bold=True, size=14)
    for cell in read_me[3]:
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    read_me.column_dimensions["A"].width = 24
    read_me.column_dimensions["B"].width = 112

    ledger = workbook.create_sheet("Candidate Input Ledger")
    ledger.append(
        ["Validated synthetic browser draft. Candidate-only; values are not source authority."]
    )
    ledger.append([])
    ledger.append(
        ["#", "Phase", "Task", "Role", "Hours", "Hourly Rate", "Fees", "Expenses", "Line Total"]
    )
    source_by_line = {
        line_number: line for line_number, line in enumerate(_source_lines(proposal), start=1)
    }
    for candidate in package["lines"]:
        source = source_by_line[candidate["lineNumber"]]
        ledger.append(
            [
                candidate["lineNumber"],
                source["phase_id"],
                source["task_id"],
                source["staffing_role"],
                candidate["estimatedHours"],
                candidate["hourlyRate"],
                candidate["estimatedFees"],
                candidate["estimatedExpenses"],
                candidate["lineTotal"],
            ]
        )
    ledger.append([])
    ledger.append(["", "", "", "Contingency", "", "", "", "", package["fixed_contingency_amount"]])
    ledger.append(["", "", "", "Candidate Total", "", "", "", "", package["draft_total"]])
    for cell in ledger[3]:
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.font = Font(bold=True)
    for column in ("F", "G", "H", "I"):
        for cell in ledger[column][3:]:
            cell.number_format = "$#,##0.00"
    for column, width in {
        "A": 6,
        "B": 12,
        "C": 12,
        "D": 22,
        "E": 12,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
    }.items():
        ledger.column_dimensions[column].width = width

    checks = workbook.create_sheet("Validation")
    checks.append(["Validation checks are recorded in the companion JSON report."])
    checks.append([])
    checks.append(["Boundary", "Value"])
    checks.append(["Data origin", "synthetic"])
    checks.append(["Candidate only", True])
    checks.append(["External writes", False])
    checks.append(["Budget submission authorized", False])
    for cell in checks[3]:
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.font = Font(bold=True)
    checks.column_dimensions["A"].width = 34
    checks.column_dimensions["B"].width = 72

    workbook.properties.title = "Synthetic Budget Sandbox Candidate"
    workbook.properties.subject = "Local candidate-only synthetic evidence"
    workbook.properties.creator = "LawFirm-os-intake"
    workbook.properties.lastModifiedBy = "LawFirm-os-intake"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def run_synthetic_budget_sandbox_xlsx_export(
    *,
    package_path: str | Path,
    repo_root: str | Path,
    out_dir: str | Path,
    generated_at: str | None = None,
) -> tuple[dict[str, Any], Path]:
    root = Path(repo_root)
    source_path = root / BUDGET_PROPOSAL_REF
    package_file = Path(package_path)
    proposal = load_json(source_path)
    proposal_sha256 = digest_text(source_path.read_text(encoding="utf-8"))
    try:
        package_text = package_file.read_text(encoding="utf-8")
        package = json.loads(package_text)
        package_parse_check: dict[str, Any] | None = None
    except (OSError, json.JSONDecodeError) as exc:
        package_text = ""
        package = None
        package_parse_check = _check(
            "package_json_parseable",
            False,
            f"Candidate package cannot be read as JSON: {exc}",
        )
    package_sha256 = digest_text(package_text)
    checks = _validate_package(package, proposal, proposal_sha256)
    if package_parse_check is not None:
        checks.insert(0, package_parse_check)
    failed = [check for check in checks if check["status"] == "failed"]
    status = (
        "synthetic_budget_sandbox_xlsx_ready_for_review"
        if not failed
        else "blocked_by_synthetic_budget_sandbox_xlsx"
    )
    report = {
        "schema_version": "0.1",
        "synthetic_budget_sandbox_xlsx_export_id": "synsandboxxlsx-"
        + digest_json({"package": package_sha256, "source": proposal_sha256}).removeprefix(
            "sha256:"
        )[:16],
        "status": status,
        "methodology_version": METHODOLOGY_VERSION,
        "source_budget_proposal_ref": BUDGET_PROPOSAL_REF,
        "source_budget_proposal_sha256": proposal_sha256,
        "candidate_package_sha256": package_sha256,
        "candidate_package_filename": package_file.name,
        "pinned_total": package.get("pinned_total") if isinstance(package, dict) else None,
        "draft_total": package.get("draft_total") if isinstance(package, dict) else None,
        "delta": package.get("delta") if isinstance(package, dict) else None,
        "line_count": len(package.get("lines", []))
        if isinstance(package, dict) and isinstance(package.get("lines"), list)
        else 0,
        "checks": checks,
        "failed_check_count": len(failed),
        "workbook_filename": SANDBOX_EXPORT_WORKBOOK_FILENAME,
        "data_origin": "synthetic",
        "candidate_only": True,
        "non_authoritative": True,
        "local_output_only": True,
        "source_mutation_performed": False,
        "external_writes_performed": False,
        "lake_write_performed": False,
        "sqlite_write_performed": False,
        "budget_submission_authorized": False,
        "matter_opening_authorized": False,
        "silent_learning_performed": False,
        "generated_at": generated_at or now_iso(),
    }
    run_dir = Path(out_dir)
    run_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = run_dir / SANDBOX_EXPORT_WORKBOOK_FILENAME
    write_json(run_dir / SANDBOX_EXPORT_REPORT_FILENAME, report)
    if not failed and isinstance(package, dict):
        _write_workbook(
            workbook_path,
            package=package,
            proposal=proposal,
            proposal_sha256=proposal_sha256,
            package_sha256=package_sha256,
        )
    elif workbook_path.exists():
        workbook_path.unlink()
    return report, run_dir
