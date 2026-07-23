"""Exporter plugin seam (structured model -> renderer) + firm-Excel exporter.

The structured budget model is the source of truth; Excel is exporter #1, not the
tool. The firm-Excel renderer matches the sanitized template shape (UTBMS
phase/task rows; Original / Billed / Remaining / New columns) but writes CORRECT
phase-subtotal and grand-total formulas and documents where it deviates from the
template's own defects. All exports are candidate-only, synthetic-only, and never
a client submission.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Protocol, runtime_checkable
from zipfile import ZipFile

from .models import (
    BudgetProposal,
    FirmExcelBudgetExport,
    FirmExcelBudgetPhase,
    FirmExcelBudgetTaskRow,
)
from .util import digest_json, load_json

FIRM_EXCEL_FORMAT_ID = "firm_excel_v0"

# Sanitized-template column positions (1-indexed): Phase/Task, Original, Billed,
# Remaining, New.
_COL_LABEL = 1
_COL_ORIGINAL = 7  # G
_COL_BILLED = 10  # J
_COL_REMAINING = 13  # M
_COL_NEW = 16  # P
_HEADER_ROW = 5

# The two documented deviations from the sanitized firm template.
_TEMPLATE_DEVIATIONS = (
    "Corrected the template's missing phase-subtotal formulas: every phase Original "
    "and New subtotal (the template only filled G17; G33/G47/G61/G77/G85 were blank) "
    "is written as an explicit SUM over that phase's task rows.",
    "Corrected the template's P85 grand-expense formula, which double-counted P129 "
    "(=...+P129+P129+P131+P133); each task row is summed exactly once.",
)


@dataclass(frozen=True)
class BudgetExportResult:
    format_id: str
    path: str
    original_total_minor_units: int
    billed_total_minor_units: int
    new_total_minor_units: int
    candidate_only: bool = True
    non_authoritative: bool = True
    not_authorized_for_client_submission: bool = True
    external_writes_performed: bool = False


@runtime_checkable
class BudgetExporter(Protocol):
    format_id: str

    def export(
        self, export_model: FirmExcelBudgetExport, path: str | Path
    ) -> BudgetExportResult: ...

    def documented_deviations(self) -> list[str]: ...


_EXPORTER_REGISTRY: dict[str, BudgetExporter] = {}


def register_budget_exporter(exporter: BudgetExporter) -> None:
    _EXPORTER_REGISTRY[exporter.format_id] = exporter


def get_budget_exporter(format_id: str) -> BudgetExporter:
    if format_id not in _EXPORTER_REGISTRY:
        raise KeyError(f"no budget exporter registered for format {format_id!r}")
    return _EXPORTER_REGISTRY[format_id]


def list_budget_exporters() -> list[str]:
    return sorted(_EXPORTER_REGISTRY)


def _dollars(minor_units: int) -> float:
    return round(minor_units / 100, 2)


def _safe_text(value: object) -> str:
    text = str(value if value is not None else "")
    return "'" + text if text.lstrip().startswith(("=", "+", "-", "@")) else text


class FirmExcelExporter:
    format_id = FIRM_EXCEL_FORMAT_ID

    def documented_deviations(self) -> list[str]:
        return list(_TEMPLATE_DEVIATIONS)

    def export(self, export_model: FirmExcelBudgetExport, path: str | Path) -> BudgetExportResult:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        target = Path(path)
        book = Workbook()
        sheet = book.active
        sheet.title = "Budget Form"

        sheet.cell(row=1, column=_COL_LABEL, value=_safe_text("BUDGET FORM"))
        sheet.cell(
            row=2,
            column=_COL_LABEL,
            value=_safe_text(
                "SYNTHETIC CANDIDATE — not a real client budget, rate authority, or submission."
            ),
        )
        sheet.cell(row=3, column=_COL_LABEL, value=_safe_text("Matter"))
        sheet.cell(row=3, column=2, value=_safe_text(export_model.matter_label))

        header = {
            _COL_LABEL: "Phase / Task",
            _COL_ORIGINAL: "Original Budgeted Amount",
            _COL_BILLED: "Amount Billed to Date",
            _COL_REMAINING: "Original Budget Amount Remaining",
            _COL_NEW: "New Budgeted Amount",
        }
        for column, label in header.items():
            cell = sheet.cell(row=_HEADER_ROW, column=column, value=_safe_text(label))
            cell.font = Font(bold=True)

        row = _HEADER_ROW + 1
        phase_rows: list[int] = []
        column_letters = {_COL_ORIGINAL: "G", _COL_BILLED: "J", _COL_NEW: "P"}
        for phase in export_model.phases:
            phase_row = row
            phase_rows.append(phase_row)
            sheet.cell(
                row=phase_row,
                column=_COL_LABEL,
                value=_safe_text(f"({phase.utbms_phase_code}) {phase.phase_name}"),
            )
            row += 1
            first_task_row = row
            for task in phase.tasks:
                sheet.cell(
                    row=row,
                    column=_COL_LABEL,
                    value=_safe_text(f"     ({task.utbms_task_code}) {task.task_name}"),
                )
                sheet.cell(
                    row=row, column=_COL_ORIGINAL, value=_dollars(task.original_amount_minor_units)
                )
                sheet.cell(
                    row=row, column=_COL_BILLED, value=_dollars(task.billed_amount_minor_units)
                )
                sheet.cell(
                    row=row,
                    column=_COL_REMAINING,
                    value=f"=G{row}-J{row}",
                )
                sheet.cell(row=row, column=_COL_NEW, value=_dollars(task.new_amount_minor_units))
                row += 1
            last_task_row = row - 1
            # CORRECTED phase subtotals: explicit SUM over this phase's task rows for
            # every phase (fixes the template's missing G33-G85) and each task summed
            # exactly once (fixes the P85 double-count of P129).
            if phase.tasks:
                for column, letter in column_letters.items():
                    sheet.cell(
                        row=phase_row,
                        column=column,
                        value=f"=SUM({letter}{first_task_row}:{letter}{last_task_row})",
                    )
            else:
                for column in column_letters:
                    sheet.cell(row=phase_row, column=column, value=0)
            sheet.cell(
                row=phase_row,
                column=_COL_REMAINING,
                value=f"=G{phase_row}-J{phase_row}",
            )

        # Grand totals as the sum of the (now complete) phase subtotals.
        total_row = row + 1
        sheet.cell(row=total_row, column=_COL_LABEL, value=_safe_text("Total Budgeted"))
        for column, letter in column_letters.items():
            if phase_rows:
                formula = "=" + "+".join(f"{letter}{pr}" for pr in phase_rows)
            else:
                formula = "0"
            sheet.cell(row=total_row, column=column, value=formula)
        sheet.cell(
            row=total_row,
            column=_COL_REMAINING,
            value=f"=G{total_row}-J{total_row}",
        )

        deviations_row = total_row + 2
        sheet.cell(
            row=deviations_row,
            column=_COL_LABEL,
            value=_safe_text("Documented deviations from the sanitized template:"),
        )
        for offset, note in enumerate(self.documented_deviations(), start=1):
            sheet.cell(row=deviations_row + offset, column=_COL_LABEL, value=_safe_text(note))

        target.parent.mkdir(parents=True, exist_ok=True)
        book.save(target)
        with ZipFile(target) as archive:
            forbidden = (
                "vbaProject",
                "externalLink",
                "connections.xml",
                "oleObject",
                "embeddings/",
            )
            if any(any(bad in name for bad in forbidden) for name in archive.namelist()):
                raise ValueError("generated firm-excel workbook contains prohibited content")

        return BudgetExportResult(
            format_id=self.format_id,
            path=str(target),
            original_total_minor_units=export_model.original_total_minor_units or 0,
            billed_total_minor_units=export_model.billed_total_minor_units or 0,
            new_total_minor_units=export_model.new_total_minor_units or 0,
        )


register_budget_exporter(FirmExcelExporter())


def read_firm_excel_task_totals(path: str | Path) -> dict:
    """Re-read a firm-Excel export and recompute totals from the task rows.

    Task rows carry numeric per-task dollars; phase/total rows carry formulas.
    Summing the numeric task cells reconstructs the totals independently of any
    spreadsheet formula evaluation, so a round-trip can reconcile to the model.
    """

    from openpyxl import load_workbook

    book = load_workbook(Path(path), data_only=False)
    sheet = book["Budget Form"]

    def _minor(value: object) -> int:
        return int(round(float(value) * 100)) if isinstance(value, (int, float)) else 0

    original = billed = new = 0
    phase_subtotals: dict[str, int] = {}
    current_phase: str | None = None
    for row in range(_HEADER_ROW + 1, sheet.max_row + 1):
        label = sheet.cell(row=row, column=_COL_LABEL).value
        original_cell = sheet.cell(row=row, column=_COL_ORIGINAL).value
        if isinstance(original_cell, str) and original_cell.startswith("="):
            # phase subtotal / grand total row
            if isinstance(label, str) and label.strip().startswith("(") and "SUM(" in original_cell:
                current_phase = label.strip().split(")")[0].strip("(")
                phase_subtotals.setdefault(current_phase, 0)
            continue
        if isinstance(original_cell, (int, float)):
            task_original = _minor(original_cell)
            original += task_original
            billed += _minor(sheet.cell(row=row, column=_COL_BILLED).value)
            new += _minor(sheet.cell(row=row, column=_COL_NEW).value)
            if current_phase is not None:
                phase_subtotals[current_phase] += task_original

    return {
        "original_total_minor_units": original,
        "billed_total_minor_units": billed,
        "new_total_minor_units": new,
        "phase_original_subtotals_minor_units": phase_subtotals,
    }


_DEFAULT_BUDGET_REF = "examples/synthetic/labor-employment/replay-inputs/epli-carrier-clean/legal_budget_proposal.json"


def firm_excel_export_from_budget(budget: BudgetProposal) -> FirmExcelBudgetExport:
    """Map an in-memory synthetic budget model to the firm-Excel export.

    Role/rate/hours stay internal; only dollars per UTBMS phase/task are exported.
    Every task total is exact integer minor units.
    """

    phases: dict[str, FirmExcelBudgetPhase] = {}
    task_totals: dict[tuple[str, str], int] = {}
    task_names: dict[tuple[str, str], str] = {}
    phase_names: dict[str, str] = {}
    phase_order: list[str] = []
    task_order: dict[str, list[str]] = {}

    for line in budget.lines:
        phase_id = line.phase_id
        task_id = line.task_id
        phase_names.setdefault(phase_id, line.phase_name)
        if phase_id not in phase_order:
            phase_order.append(phase_id)
            task_order[phase_id] = []
        key = (phase_id, task_id)
        if task_id not in task_order[phase_id]:
            task_order[phase_id].append(task_id)
        task_names.setdefault(key, line.task_name)
        fees_minor = int(round((line.estimated_fees or 0.0) * 100))
        expenses_minor = int(round(line.estimated_expenses * 100))
        task_totals[key] = task_totals.get(key, 0) + fees_minor + expenses_minor

    for phase_id in phase_order:
        phases[phase_id] = FirmExcelBudgetPhase(
            utbms_phase_code=phase_id,
            phase_name=phase_names[phase_id],
            tasks=[
                FirmExcelBudgetTaskRow(
                    utbms_task_code=task_id,
                    task_name=task_names[(phase_id, task_id)],
                    original_amount_minor_units=task_totals[(phase_id, task_id)],
                    billed_amount_minor_units=0,
                    new_amount_minor_units=task_totals[(phase_id, task_id)],
                )
                for task_id in task_order[phase_id]
            ],
        )

    basis = {"budget_proposal_id": budget.budget_proposal_id, "phases": phase_order}
    return FirmExcelBudgetExport(
        export_id="firm-excel-export-" + digest_json(basis).removeprefix("sha256:")[:16],
        phases=[phases[phase_id] for phase_id in phase_order],
        documented_deviations=list(_TEMPLATE_DEVIATIONS),
    )


def firm_excel_export_from_projection_report(
    *, repo_root: str | Path, budget_ref: str = _DEFAULT_BUDGET_REF
) -> FirmExcelBudgetExport:
    """Load a serialized synthetic budget model and map it to the firm-Excel export."""

    root = Path(repo_root)
    budget = BudgetProposal.model_validate(load_json(root / budget_ref))
    return firm_excel_export_from_budget(budget)
