"""Build a governed, synthetic-only guideline projection review artifact.

The report intentionally distinguishes a proposal from a synthetic guideline
projection.  It is a frozen, local counterfactual proof, not a carrier decision,
rate authorization, submission, or calibration surface.
"""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile

from .confirmation import bind_confirmation_to_packet_evidence
from .context import load_profile
from .guidelines import build_carrier_compliant_projection, build_carrier_preapproval_report
from .models import (
    HumanConfirmation,
    SyntheticGuidelineProjectionWorkbenchCheck,
    SyntheticGuidelineProjectionWorkbenchReport,
    SyntheticGuidelineProjectionWorkbenchSource,
    SyntheticGuidelineProjectionWorkbenchView,
)
from .rates import load_rate_card, resolve_role_rates
from .util import digest_json, digest_text, load_json, now_iso, write_json
from .workflow import run_budget, run_preflight

SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_REPORT_FILENAME = (
    "synthetic_guideline_projection_workbench_report.json"
)
SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_MARKDOWN_FILENAME = (
    "synthetic_guideline_projection_workbench.md"
)
SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_WORKBOOK_FILENAME = (
    "synthetic_guideline_projection_workbench.xlsx"
)
METHODOLOGY_VERSION = "synthetic_guideline_projection_workbench.v0_1"

SOURCE_REFS = (
    (
        "inbound",
        "Synthetic medical-malpractice assignment",
        "examples/synthetic/inbound/carrier-assignment-medmal.json",
    ),
    (
        "confirmation",
        "Synthetic human confirmation template",
        "examples/synthetic/confirmations/carrier-assignment-medmal.confirmation-template.json",
    ),
    (
        "profile",
        "Synthetic insurance-defense profile",
        "context/synthetic-profiles/insurance-defense.yaml",
    ),
    ("guideline", "Synthetic candidate guideline", "config/synthetic-carrier-guideline.yaml"),
    ("rate_card", "Synthetic candidate rate card", "config/synthetic-carrier-rate-card.yaml"),
)
REQUIRED_THRESHOLDS = {
    "experts_over_count",
    "expert_spend_over_amount",
    "depositions_over_count",
    "research_hours_over",
    "vendor_spend_over_amount",
}


def _check(
    check_id: str, passed: bool, message: str, *refs: str
) -> SyntheticGuidelineProjectionWorkbenchCheck:
    return SyntheticGuidelineProjectionWorkbenchCheck(
        check_id=check_id,
        status="passed" if passed else "failed",
        message=message,
        evidence_refs=list(refs),
    )


def _source_manifest(root: Path) -> list[SyntheticGuidelineProjectionWorkbenchSource]:
    return [
        SyntheticGuidelineProjectionWorkbenchSource(
            source_id=source_id,
            label=label,
            source_ref=source_ref,
            source_sha256=digest_text((root / source_ref).read_text(encoding="utf-8")),
            used_for="frozen synthetic projection replay",
        )
        for source_id, label, source_ref in SOURCE_REFS
    ]


def _carrier_confirmation(raw: dict, carrier_name: str, aliases: list[str]) -> dict:
    copied = deepcopy(raw)
    for party in copied["confirmed_parties"]:
        if party["confirmed_role"] in {"insurance_carrier", "payer"}:
            party["name"] = carrier_name
            party["aliases"] = aliases
    return copied


def _thresholds_complete(guideline: dict, carrier_id: str) -> bool:
    thresholds = guideline.get("carriers", {}).get(carrier_id, {}).get("pre_approval_thresholds")
    if not isinstance(thresholds, dict) or set(thresholds) != REQUIRED_THRESHOLDS:
        return False
    try:
        return all(float(thresholds[key]) >= 0 for key in REQUIRED_THRESHOLDS)
    except (TypeError, ValueError):
        return False


def _view(
    *,
    budget,
    guideline: dict,
    rate_card: dict,
    profile: dict,
    confirmation: HumanConfirmation,
    carrier_id: str,
    generated_at: str,
) -> SyntheticGuidelineProjectionWorkbenchView:
    rate_resolution = resolve_role_rates(
        profile=profile,
        confirmation=confirmation,
        rate_card=rate_card,
    )
    projection = build_carrier_compliant_projection(
        budget,
        guideline=guideline,
        guideline_ref="config/synthetic-carrier-guideline.yaml",
        carrier_id=carrier_id,
        rate_resolution=rate_resolution,
    )
    preapproval = build_carrier_preapproval_report(
        budget,
        guideline=guideline,
        guideline_ref="config/synthetic-carrier-guideline.yaml",
        carrier_id=carrier_id,
    )
    if projection is None or preapproval is None:
        raise ValueError(
            "synthetic guideline fixture must resolve projection and preapproval report"
        )
    projection = projection.model_copy(
        update={"projection_id": f"synthetic-guideline-projection-{carrier_id}.v0_1"}
    )
    preapproval = preapproval.model_copy(
        update={
            "report_id": f"synthetic-guideline-preapproval-{carrier_id}.v0_1",
            "requirements": [
                item.model_copy(
                    update={
                        "requirement_id": (
                            f"synthetic-guideline-preapproval-{carrier_id}-{item.threshold_id}.v0_1"
                        )
                    }
                )
                for item in preapproval.requirements
            ],
            "generated_at": generated_at,
        }
    )
    return SyntheticGuidelineProjectionWorkbenchView(
        carrier_id=carrier_id,
        carrier_name=str(guideline["carriers"][carrier_id]["carrier_name"]),
        state=rate_resolution.state,
        rate_card_effective_date=rate_resolution.effective_date,
        projection=projection,
        preapproval_report=preapproval,
        rate_resolution=rate_resolution.model_dump(mode="json"),
        gross_reductions=projection.total_delta,
        gross_increases=projection.compliant_increase_amount,
        net_delta=projection.total_delta_signed,
    )


def build_synthetic_guideline_projection_workbench_report(
    *, repo_root: str | Path, generated_at: str | None = None
) -> SyntheticGuidelineProjectionWorkbenchReport:
    root = Path(repo_root)
    report_generated_at = generated_at or now_iso()
    inbound = root / SOURCE_REFS[0][2]
    confirmation_path = root / SOURCE_REFS[1][2]
    profile = load_profile(root / SOURCE_REFS[2][2])
    from .guidelines import load_carrier_guideline

    guideline = load_carrier_guideline(root / SOURCE_REFS[3][2])
    rate_card = load_rate_card(root / SOURCE_REFS[4][2])
    with TemporaryDirectory(prefix="lawfirm-os-intake-projection-") as temporary:
        replay_root = Path(temporary)
        packet, preflight_dir = run_preflight(
            inbound, root / SOURCE_REFS[2][2], replay_root / "preflight"
        )
        raw_confirmation = load_json(confirmation_path)
        raw_confirmation["preflight_packet_id"] = packet.packet_id
        harbor_confirmation = bind_confirmation_to_packet_evidence(
            packet, HumanConfirmation.model_validate(raw_confirmation)
        )
        harbor_confirmation_path = replay_root / "harbor-confirmation.json"
        harbor_confirmation_path.write_text(harbor_confirmation.model_dump_json(), encoding="utf-8")
        budget, _ = run_budget(
            preflight_dir / "intake_preflight_packet.json",
            harbor_confirmation_path,
            root / SOURCE_REFS[2][2],
            replay_root / "budget",
        )
    # Keep the reusable projection proof free of run-specific output identities.
    budget = budget.model_copy(
        update={
            "budget_proposal_id": "synthetic-guideline-projection-medmal.v0_1",
            "preflight_packet_id": "synthetic-guideline-projection-preflight.v0_1",
            "confirmation_id": "synthetic-guideline-projection-confirmation.v0_1",
            "carrier_compliant_projection": None,
            "carrier_preapproval_report": None,
        }
    )

    cascade_raw = _carrier_confirmation(raw_confirmation, "Cascade Mutual", ["Cascade"])
    cascade_confirmation = bind_confirmation_to_packet_evidence(
        packet, HumanConfirmation.model_validate(cascade_raw)
    )
    views = [
        _view(
            budget=budget,
            guideline=guideline,
            rate_card=rate_card,
            profile=profile,
            confirmation=harbor_confirmation,
            carrier_id="synthetic-carrier-a",
            generated_at=report_generated_at,
        ),
        _view(
            budget=budget,
            guideline=guideline,
            rate_card=rate_card,
            profile=profile,
            confirmation=cascade_confirmation,
            carrier_id="synthetic-carrier-b",
            generated_at=report_generated_at,
        ),
    ]
    proposal_hash = digest_json(
        {
            "budget_proposal_id": budget.budget_proposal_id,
            "currency": budget.currency,
            "subtotal_fees": budget.subtotal_fees,
            "subtotal_expenses": budget.subtotal_expenses,
            "contingency_amount": budget.contingency_amount,
            "total_proposed_budget": budget.total_proposed_budget,
            "lines": [
                {
                    "phase_id": line.phase_id,
                    "task_id": line.task_id,
                    "staffing_role": line.staffing_role,
                    "estimated_hours": line.estimated_hours,
                    "hourly_rate": line.hourly_rate,
                    "estimated_fees": line.estimated_fees,
                    "estimated_expenses": line.estimated_expenses,
                }
                for line in budget.lines
            ],
        }
    )
    checks = [
        _check(
            "synthetic_sources_present_and_hashed",
            True,
            "Every frozen source has a local path and SHA-256 digest.",
            *[item[2] for item in SOURCE_REFS],
        ),
        _check(
            "proposal_preserved_across_counterfactuals",
            all(view.projection.proposed_total == budget.total_proposed_budget for view in views),
            "Both synthetic guideline views must begin with the same immutable proposal total.",
        ),
        _check(
            "rate_resolution_is_synthetic_and_priced",
            all(
                view.rate_resolution["source"] == "carrier_rate_card"
                and not view.rate_resolution["review_required"]
                and view.projection.projection_pricing_status == "priced"
                for view in views
            ),
            "A staffing reshape must bind to a synthetic carrier/state/title schedule or block.",
            SOURCE_REFS[4][2],
        ),
        _check(
            "thresholds_complete_and_evaluable",
            all(_thresholds_complete(guideline, view.carrier_id) for view in views)
            and all(
                len(view.preapproval_report.requirements) == len(REQUIRED_THRESHOLDS)
                and all(item.status != "unknown" for item in view.preapproval_report.requirements)
                for view in views
            ),
            "All declared preapproval rules must be numeric, complete, and evaluable; unknown blocks readiness.",
            SOURCE_REFS[3][2],
        ),
        _check(
            "signed_delta_partition_reconciles",
            all(
                round(view.gross_reductions - view.gross_increases, 2) == view.net_delta
                and round(
                    view.projection.proposed_total - (view.projection.compliant_total or 0), 2
                )
                == view.net_delta
                for view in views
            ),
            "Gross reductions, gross increases, signed net delta, and projected total must reconcile.",
        ),
        _check(
            "line_and_leverage_reconcile",
            all(
                round(
                    sum(line.compliant_line_total or 0 for line in view.projection.lines)
                    + float(view.projection.compliant_contingency_amount or 0),
                    2,
                )
                == view.projection.compliant_total
                and abs(
                    sum(item.compliant_hours_percent for item in view.projection.leverage_summary)
                    - 100.0
                )
                <= 0.02
                for view in views
            ),
            "Projected lines and leverage percentages must reconcile before display.",
        ),
        _check(
            "candidate_only_non_submission_boundary",
            all(
                view.projection.not_authorized_for_client_submission
                and not view.projection.external_writes_performed
                and not view.preapproval_report.carrier_submission_authorized
                for view in views
            ),
            "This artifact is synthetic review evidence, never a carrier approval or submission.",
        ),
    ]
    failed_count = sum(check.status == "failed" for check in checks)
    report_basis = {
        "proposal_sha256": proposal_hash,
        "views": [view.model_dump(mode="json") for view in views],
    }
    return SyntheticGuidelineProjectionWorkbenchReport(
        synthetic_guideline_projection_workbench_report_id=(
            "synguidelineprojection-" + digest_json(report_basis).removeprefix("sha256:")[:16]
        ),
        status=(
            "synthetic_guideline_projection_workbench_ready_for_review"
            if not failed_count
            else "blocked_by_synthetic_guideline_projection_workbench"
        ),
        budget_proposal_id=budget.budget_proposal_id,
        budget_proposal_sha256=proposal_hash,
        currency=budget.currency,
        proposal_total=budget.total_proposed_budget,
        views=views,
        source_manifest=_source_manifest(root),
        checks=checks,
        failed_check_count=failed_count,
        workbook_filename=SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_WORKBOOK_FILENAME,
        markdown_filename=SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_MARKDOWN_FILENAME,
        display_banner={
            "summary": "Synthetic guideline projection only. Proposed budget is preserved; projected values are not carrier approval, authorization, or submission.",
            "candidate_only": True,
            "synthetic_only": True,
            "read_only_ui": True,
            "blocked_actions": [
                "guideline_editing",
                "carrier_approval",
                "budget_submission",
                "rate_calibration",
                "exception_lake_write",
                "sqlite_write",
            ],
        },
        candidate_exception_lake_labels=[
            "synthetic_guideline_projection_review_candidate",
            "synthetic_preapproval_threshold_gap_candidate",
        ],
        required_next_gates=[
            "human_review_before_real_guideline_use",
            "legal_knowledge_runtime_governed_rate_research",
            "orchestrator_owned_submission_contract",
        ],
        generated_at=report_generated_at,
    )


def _safe_text(value: object) -> str:
    text = str(value if value is not None else "")
    return "'" + text if text.lstrip().startswith(("=", "+", "-", "@")) else text


def _write_workbook(report: SyntheticGuidelineProjectionWorkbenchReport, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    book = Workbook()
    readme = book.active
    readme.title = "Read Me"
    for row in [
        ["Synthetic Guideline Projection Workbench"],
        ["Status", report.status],
        ["Proposal hash", report.budget_proposal_sha256],
        ["Boundary", report.display_banner["summary"]],
        ["Workbook formulas", "none"],
    ]:
        readme.append([_safe_text(value) for value in row])
    for view in report.views:
        sheet = book.create_sheet(view.carrier_id[-1:].upper() + " Projection")
        sheet.append(
            [_safe_text("Synthetic candidate projection; not carrier approved or submittable.")]
        )
        sheet.append([])
        sheet.append(
            [
                "Phase",
                "Task",
                "Role",
                "Projected Role",
                "Hours",
                "Proposed",
                "Projected",
                "Net Delta",
                "Reason",
            ]
        )
        for line in view.projection.lines:
            sheet.append(
                [
                    _safe_text(value)
                    for value in [
                        line.phase_id,
                        line.task_id,
                        line.staffing_role,
                        line.compliant_staffing_role,
                        line.proposed_hours,
                        line.proposed_line_total,
                        line.compliant_line_total,
                        line.line_delta_signed,
                        line.note,
                    ]
                ]
            )
        fill = PatternFill("solid", fgColor="234157")
        for cell in sheet[3]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
        sheet.freeze_panes = "A4"
        for column, width in {
            "A": 14,
            "B": 14,
            "C": 22,
            "D": 22,
            "E": 12,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 88,
        }.items():
            sheet.column_dimensions[column].width = width
        for column in ("F", "G", "H"):
            for cell in sheet[column][3:]:
                cell.number_format = "$#,##0.00"
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    with ZipFile(path) as archive:
        forbidden = ("vbaProject", "externalLinks/", "connections.xml", "embeddings/", "oleObject")
        if any(any(marker in name for marker in forbidden) for name in archive.namelist()):
            raise ValueError(
                "generated projection workbook contains prohibited active/external content"
            )


def _write_markdown(report: SyntheticGuidelineProjectionWorkbenchReport, path: Path) -> None:
    rows = [
        "# Synthetic Guideline Projection Workbench",
        "",
        report.display_banner["summary"],
        "",
        f"- Proposal hash: `{report.budget_proposal_sha256}`",
        f"- Proposal total: `{report.proposal_total}` {report.currency}",
        "",
        "## Synthetic Counterfactuals",
        "",
    ]
    rows.extend(
        f"- `{view.carrier_name}`: projected `{view.projection.compliant_total}`, gross reductions `{view.gross_reductions}`, gross increases `{view.gross_increases}`, net delta `{view.net_delta}`"
        for view in report.views
    )
    rows.extend(["", "## Checks", ""])
    rows.extend(
        f"- `{check.status}` `{check.check_id}`: {check.message}" for check in report.checks
    )
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def run_synthetic_guideline_projection_workbench(
    *, repo_root: str | Path, out_dir: str | Path, generated_at: str | None = None
) -> tuple[SyntheticGuidelineProjectionWorkbenchReport, Path]:
    report = build_synthetic_guideline_projection_workbench_report(
        repo_root=repo_root, generated_at=generated_at
    )
    run_dir = Path(out_dir)
    run_dir.mkdir(parents=True, exist_ok=True)
    write_json(
        run_dir / SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_REPORT_FILENAME,
        report.model_dump(mode="json"),
    )
    _write_markdown(report, run_dir / SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_MARKDOWN_FILENAME)
    if report.status == "synthetic_guideline_projection_workbench_ready_for_review":
        _write_workbook(
            report, run_dir / SYNTHETIC_GUIDELINE_PROJECTION_WORKBENCH_WORKBOOK_FILENAME
        )
    return report, run_dir
