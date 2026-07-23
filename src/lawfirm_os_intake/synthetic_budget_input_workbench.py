"""Build a deterministic, synthetic-only budget-input lineage ledger.

The generated XLSX is a review/export artifact.  It never accepts browser edits
or combines rate-card, guideline, benchmark, or actuals context into the fixed
proposal math.
"""

from __future__ import annotations

from pathlib import Path

from .models import (
    BudgetProposal,
    SyntheticBudgetInputWorkbenchCheck,
    SyntheticBudgetInputWorkbenchContextLane,
    SyntheticBudgetInputWorkbenchLine,
    SyntheticBudgetInputWorkbenchReport,
)
from .util import digest_json, digest_text, load_json, now_iso, write_json

SYNTHETIC_BUDGET_INPUT_WORKBENCH_REPORT_FILENAME = "synthetic_budget_input_workbench_report.json"
SYNTHETIC_BUDGET_INPUT_WORKBENCH_MARKDOWN_FILENAME = "synthetic_budget_input_workbench.md"
SYNTHETIC_BUDGET_INPUT_WORKBOOK_FILENAME = "synthetic_budget_input_workbench.xlsx"
METHODOLOGY_VERSION = "synthetic_budget_input_workbench.v0_1"
BUDGET_PROPOSAL_REF = (
    "examples/synthetic/labor-employment/replay-inputs/epli-carrier-clean/"
    "legal_budget_proposal.json"
)
CONTEXT_SOURCES = (
    (
        "synthetic_rate_card",
        "Synthetic rate card",
        "config/synthetic-carrier-rate-card.yaml",
        "This fixed proposal declares synthetic_profile line rates; the separate synthetic rate card is not a pricing input.",
    ),
    (
        "synthetic_carrier_guideline",
        "Synthetic carrier guideline",
        "config/synthetic-carrier-guideline.yaml",
        "Guidelines remain a separate candidate policy lane and are not applied to this fixed proposal.",
    ),
    (
        "synthetic_actuals",
        "Synthetic actuals",
        "examples/synthetic/labor-employment/replay-inputs/epli-carrier-clean/budget_actuals_source.json",
        "Actuals are a post-budget variance lane and are never added to proposal inputs.",
    ),
    (
        "synthetic_benchmark",
        "Synthetic rate benchmark",
        "examples/synthetic/benchmarks/synthetic-rate-benchmark-snapshot.json",
        "Benchmark evidence is candidate context only and is not a rate, multiplier, or cap for this proposal.",
    ),
)


def _check(
    check_id: str, passed: bool, message: str, *refs: str
) -> SyntheticBudgetInputWorkbenchCheck:
    return SyntheticBudgetInputWorkbenchCheck(
        check_id=check_id,
        status="passed" if passed else "failed",
        message=message,
        evidence_refs=[ref for ref in refs if ref],
    )


def _line_total(fees: float | None, expenses: float) -> float:
    return round(float(fees or 0) + float(expenses), 2)


def _context_lanes(root: Path) -> list[SyntheticBudgetInputWorkbenchContextLane]:
    lanes = [
        SyntheticBudgetInputWorkbenchContextLane(
            lane_id="budget_proposal",
            label="Fixed synthetic budget proposal",
            inclusion="used_for_budget_math",
            reason="All displayed line math comes only from the pinned synthetic proposal.",
            source_ref=BUDGET_PROPOSAL_REF,
            source_sha256=digest_text((root / BUDGET_PROPOSAL_REF).read_text(encoding="utf-8")),
        )
    ]
    for lane_id, label, source_ref, reason in CONTEXT_SOURCES:
        path = root / source_ref
        lanes.append(
            SyntheticBudgetInputWorkbenchContextLane(
                lane_id=lane_id,
                label=label,
                inclusion="excluded_context_only",
                reason=reason,
                source_ref=source_ref if path.is_file() else None,
                source_sha256=(
                    digest_text(path.read_text(encoding="utf-8")) if path.is_file() else None
                ),
            )
        )
    return lanes


def build_synthetic_budget_input_workbench_report(
    *,
    repo_root: str | Path,
    generated_at: str | None = None,
    budget_ref: str = BUDGET_PROPOSAL_REF,
) -> SyntheticBudgetInputWorkbenchReport:
    root = Path(repo_root)
    return _build_synthetic_budget_input_workbench_report(
        budget_path=root / budget_ref,
        repo_root=root,
        generated_at=generated_at,
        budget_ref=budget_ref,
    )


def _build_synthetic_budget_input_workbench_report(
    *,
    budget_path: Path,
    repo_root: Path,
    generated_at: str | None = None,
    budget_ref: str = BUDGET_PROPOSAL_REF,
) -> SyntheticBudgetInputWorkbenchReport:
    """Build from a supplied proposal path.

    ``budget_ref`` selects which governed synthetic proposal family is exposed;
    it defaults to the pinned EPLI proposal. Also used by hostile-fixture tests.
    """

    source_paths = [
        budget_path,
        *[repo_root / source_ref for _, _, source_ref, _ in CONTEXT_SOURCES],
    ]
    source_bytes_before = {path: path.read_bytes() for path in source_paths if path.is_file()}
    budget_text = budget_path.read_text(encoding="utf-8")
    budget = BudgetProposal.model_validate(load_json(budget_path))
    lines = [
        SyntheticBudgetInputWorkbenchLine(
            line_number=index,
            phase_id=line.phase_id,
            phase_name=line.phase_name,
            task_id=line.task_id,
            task_name=line.task_name,
            staffing_role=line.staffing_role,
            estimated_hours=line.estimated_hours,
            hourly_rate=line.hourly_rate,
            rate_source=line.rate_source,
            rate_is_synthetic=line.rate_is_synthetic,
            estimated_fees=line.estimated_fees,
            estimated_expenses=line.estimated_expenses,
            line_total=_line_total(line.estimated_fees, line.estimated_expenses),
            calculation_formula=line.calculation_formula,
            estimate_basis=line.estimate_basis,
            estimate_basis_refs=line.estimate_basis_refs,
        )
        for index, line in enumerate(budget.lines, start=1)
    ]
    fee_total = round(sum(float(line.estimated_fees or 0) for line in lines), 2)
    expense_total = round(sum(line.estimated_expenses for line in lines), 2)
    total = round(fee_total + expense_total + float(budget.contingency_amount or 0), 2)
    line_fee_math = all(
        line.hourly_rate is None
        or line.estimated_fees is None
        or round(line.estimated_hours * line.hourly_rate, 2) == line.estimated_fees
        for line in lines
    )
    line_total_math = all(
        line.line_total == _line_total(line.estimated_fees, line.estimated_expenses)
        for line in lines
    )
    source_inputs_unchanged = all(
        path.is_file() and path.read_bytes() == content
        for path, content in source_bytes_before.items()
    )
    checks = [
        _check(
            "synthetic_proposal_boundary",
            budget.display_banner.get("data_scope") == "synthetic_only"
            and budget.display_banner.get("candidate_only") is True
            and budget.not_authorized_for_client_submission is True,
            "The pinned proposal must declare synthetic candidate-only and non-submittable status.",
            budget_ref,
        ),
        _check(
            "line_fee_math_reconciles",
            line_fee_math,
            "Each priced line must equal estimated hours multiplied by its displayed hourly rate.",
            budget_ref,
        ),
        _check(
            "line_total_math_reconciles",
            line_total_math,
            "Every line total must equal its displayed fees plus expenses.",
            budget_ref,
        ),
        _check(
            "proposal_totals_reconcile",
            fee_total == budget.subtotal_fees
            and expense_total == budget.subtotal_expenses
            and total == budget.total_proposed_budget,
            "Displayed line totals must reconcile exactly to proposal subtotals and total.",
            budget_ref,
        ),
        _check(
            "all_rates_synthetic",
            all(
                line.rate_is_synthetic and line.rate_source == "synthetic_profile" for line in lines
            ),
            "This fixed replay accepts only declared synthetic_profile rates.",
            budget_ref,
        ),
        _check(
            "estimate_basis_refs_present",
            all(line.estimate_basis and line.estimate_basis_refs for line in lines),
            "Every displayed budget line requires an estimate basis and at least one basis reference.",
            budget_ref,
        ),
        _check(
            "excluded_context_lanes_not_included_in_math",
            all(
                lane.inclusion == "excluded_context_only" for lane in _context_lanes(repo_root)[1:]
            ),
            "Rate card, guideline, actuals, and benchmark lanes remain explicit context only.",
            *[source_ref for _, _, source_ref, _ in CONTEXT_SOURCES],
        ),
        _check(
            "declared_context_sources_present_and_hashed",
            all(
                lane.source_ref is not None and lane.source_sha256 is not None
                for lane in _context_lanes(repo_root)
            ),
            "Every declared input or excluded context lane must resolve to a local source reference and hash.",
            budget_ref,
            *[source_ref for _, _, source_ref, _ in CONTEXT_SOURCES],
        ),
        _check(
            "source_inputs_unchanged_during_build",
            source_inputs_unchanged,
            "Pinned proposal and context inputs must remain byte-identical while the workbench builds.",
            budget_ref,
            *[source_ref for _, _, source_ref, _ in CONTEXT_SOURCES],
        ),
    ]
    failed_check_count = sum(check.status == "failed" for check in checks)
    report_basis = {
        "budget_proposal_sha256": digest_text(budget_text),
        "lines": [line.model_dump(mode="json") for line in lines],
        "methodology_version": METHODOLOGY_VERSION,
    }
    return SyntheticBudgetInputWorkbenchReport(
        synthetic_budget_input_workbench_report_id=(
            "synbudgetinputworkbench-" + digest_json(report_basis).removeprefix("sha256:")[:16]
        ),
        status=(
            "synthetic_budget_input_workbench_ready_for_review"
            if failed_check_count == 0
            else "blocked_by_synthetic_budget_input_workbench"
        ),
        budget_proposal_id=budget.budget_proposal_id,
        budget_proposal_ref=budget_ref,
        budget_proposal_sha256=digest_text(budget_text),
        preflight_packet_id=budget.preflight_packet_id,
        confirmation_id=budget.confirmation_id,
        practice_profile_id=budget.practice_profile_id,
        matter_family=budget.matter_family,
        representation_posture=budget.representation_posture,
        pricing_status=budget.pricing_status,
        currency=budget.currency,
        lines=lines,
        line_count=len(lines),
        subtotal_fees=budget.subtotal_fees,
        subtotal_expenses=budget.subtotal_expenses,
        contingency_amount=budget.contingency_amount,
        total_proposed_budget=budget.total_proposed_budget,
        context_lanes=_context_lanes(repo_root),
        checks=checks,
        failed_check_count=failed_check_count,
        workbook_filename=SYNTHETIC_BUDGET_INPUT_WORKBOOK_FILENAME,
        markdown_filename=SYNTHETIC_BUDGET_INPUT_WORKBENCH_MARKDOWN_FILENAME,
        display_banner={
            "summary": (
                "Synthetic candidate budget input ledger only. It exposes the pinned proposal's "
                "numbers and formulas; it does not accept edits, price a matter, or authorize submission."
            ),
            "candidate_only": True,
            "synthetic_only": True,
            "read_only_ui": True,
            "blocked_actions": [
                "browser_input_editing",
                "automatic_rate_resolution",
                "carrier_guideline_application",
                "actuals_calibration",
                "budget_submission",
                "exception_lake_write",
                "sqlite_write",
            ],
        },
        candidate_exception_lake_labels=[
            "synthetic_budget_input_lineage_review_candidate",
            "synthetic_budget_math_reconciliation_candidate",
        ],
        required_next_gates=[
            "orchestrator_owned_edit_session_contract_before_runtime_input",
            "human_confirmation_before_matter_specific_pricing",
            "legal_knowledge_runtime_review_before_real_rate_or_benchmark_use",
        ],
        generated_at=generated_at or now_iso(),
    )


def _style_sheet(ws, *, header_row: int) -> None:
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="234157")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"


def _write_workbook(report: SyntheticBudgetInputWorkbenchReport, path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    read_me = workbook.active
    read_me.title = "Read Me"
    for row in [
        ["Synthetic Budget Input Ledger"],
        ["Status", report.status],
        ["Proposal", report.budget_proposal_ref],
        ["Proposal hash", report.budget_proposal_sha256],
        ["Generated at", report.generated_at],
        ["Boundary", report.display_banner["summary"]],
        ["Refresh", "Edit the checked-in synthetic proposal, then regenerate this workbook."],
        ["Workbook formulas", "none; values are audited snapshots of the proposal."],
    ]:
        read_me.append(row)
    read_me.column_dimensions["A"].width = 22
    read_me.column_dimensions["B"].width = 118

    ledger = workbook.create_sheet("Input Ledger")
    ledger.append(["Synthetic candidate-only input snapshot. No browser edits or runtime pricing."])
    ledger.append([])
    ledger.append(
        [
            "#",
            "Phase",
            "Task",
            "Role",
            "Hours",
            "Hourly Rate",
            "Fees",
            "Expenses",
            "Line Total",
            "Rate Source",
            "Estimate Basis",
            "Basis References",
            "Formula",
        ]
    )
    for line in report.lines:
        ledger.append(
            [
                line.line_number,
                line.phase_id,
                line.task_id,
                line.staffing_role,
                line.estimated_hours,
                line.hourly_rate,
                line.estimated_fees,
                line.estimated_expenses,
                line.line_total,
                line.rate_source,
                line.estimate_basis,
                " | ".join(line.estimate_basis_refs),
                line.calculation_formula or "",
            ]
        )
    _style_sheet(ledger, header_row=3)
    for column in ("F", "G", "H", "I"):
        for cell in ledger[column][3:]:
            cell.number_format = "$#,##0.00"
    for column, width in {
        "A": 6,
        "B": 12,
        "C": 12,
        "D": 20,
        "E": 10,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 22,
        "K": 20,
        "L": 72,
        "M": 68,
    }.items():
        ledger.column_dimensions[column].width = width

    context = workbook.create_sheet("Context Lanes")
    context.append(
        ["Context is visible for provenance only. Excluded lanes do not affect ledger math."]
    )
    context.append([])
    context.append(["Lane", "Inclusion", "Reason", "Source", "Source Hash"])
    for lane in report.context_lanes:
        context.append(
            [
                lane.label,
                lane.inclusion,
                lane.reason,
                lane.source_ref or "not supplied",
                lane.source_sha256 or "not supplied",
            ]
        )
    _style_sheet(context, header_row=3)
    for column, width in {"A": 28, "B": 26, "C": 80, "D": 70, "E": 76}.items():
        context.column_dimensions[column].width = width

    workbook.properties.title = "Synthetic Budget Input Ledger"
    workbook.properties.subject = "Synthetic candidate-only local evidence"
    workbook.properties.creator = "LawFirm-os-intake"
    workbook.properties.lastModifiedBy = "LawFirm-os-intake"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_markdown(report: SyntheticBudgetInputWorkbenchReport, path: Path) -> None:
    lines = [
        "# Synthetic Budget Input Ledger",
        "",
        report.display_banner["summary"],
        "",
        f"- Proposal: `{report.budget_proposal_ref}` ({report.budget_proposal_sha256})",
        f"- Total: `{report.total_proposed_budget}` {report.currency}",
        f"- Lines: `{report.line_count}`",
        "",
        "## Checks",
        "",
    ]
    lines.extend(
        f"- `{check.status}` `{check.check_id}`: {check.message}" for check in report.checks
    )
    lines.extend(["", "## Context Lanes", ""])
    lines.extend(
        f"- `{lane.inclusion}` `{lane.label}`: {lane.reason}" for lane in report.context_lanes
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_synthetic_budget_input_workbench(
    *, repo_root: str | Path, out_dir: str | Path, generated_at: str | None = None
) -> tuple[SyntheticBudgetInputWorkbenchReport, Path]:
    report = build_synthetic_budget_input_workbench_report(
        repo_root=repo_root, generated_at=generated_at
    )
    run_dir = Path(out_dir)
    run_dir.mkdir(parents=True, exist_ok=True)
    write_json(
        run_dir / SYNTHETIC_BUDGET_INPUT_WORKBENCH_REPORT_FILENAME, report.model_dump(mode="json")
    )
    _write_markdown(report, run_dir / SYNTHETIC_BUDGET_INPUT_WORKBENCH_MARKDOWN_FILENAME)
    if report.status == "synthetic_budget_input_workbench_ready_for_review":
        _write_workbook(report, run_dir / SYNTHETIC_BUDGET_INPUT_WORKBOOK_FILENAME)
    return report, run_dir
