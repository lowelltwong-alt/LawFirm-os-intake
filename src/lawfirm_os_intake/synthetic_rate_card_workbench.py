"""Build a deterministic, synthetic-only rate-card workbench artifact.

This module deliberately exposes a catalog, not matter-specific pricing.  Matter
pricing continues to require a confirmed carrier, jurisdiction, and role through
``rates.resolve_role_rates``.  The workbench is local candidate evidence only.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml

from .models import (
    SyntheticRateCardWorkbenchCheck,
    SyntheticRateCardWorkbenchReport,
    SyntheticRateCardWorkbenchRow,
    SyntheticRateCardWorkbenchStateSummary,
)
from .util import digest_json, digest_text, write_json

SYNTHETIC_RATE_CARD_WORKBENCH_REPORT_FILENAME = "synthetic_rate_card_workbench_report.json"
SYNTHETIC_RATE_CARD_WORKBENCH_MARKDOWN_FILENAME = "synthetic_rate_card_workbench.md"
SYNTHETIC_RATE_CARD_WORKBOOK_FILENAME = "synthetic_rate_card_workbench.xlsx"
METHODOLOGY_VERSION = "synthetic_rate_card_workbench.v0_1"


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _relative_ref(path: Path, repo_root: Path) -> str:
    try:
        return path.resolve().relative_to(repo_root.resolve()).as_posix()
    except ValueError:
        return path.name


def _check(
    check_id: str, passed: bool, message: str, *refs: str
) -> SyntheticRateCardWorkbenchCheck:
    return SyntheticRateCardWorkbenchCheck(
        check_id=check_id,
        status="passed" if passed else "failed",
        message=message,
        evidence_refs=[ref for ref in refs if ref],
    )


def _safe_text(value: Any) -> str:
    """Prevent catalog text from being interpreted as a spreadsheet formula."""

    text = str(value)
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _truthy_real_data_flags(value: Any, prefix: str = "") -> list[str]:
    """Find any declared real-data flag that attempts to opt this catalog into real data."""

    findings: list[str] = []
    if isinstance(value, dict):
        for key, nested in value.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            if str(key).startswith("contains_real_") and nested is True:
                findings.append(path)
            findings.extend(_truthy_real_data_flags(nested, path))
    elif isinstance(value, list):
        for index, nested in enumerate(value):
            findings.extend(_truthy_real_data_flags(nested, f"{prefix}[{index}]"))
    return findings


def _as_mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _collect_rows_and_checks(
    card: dict[str, Any],
    *,
    rate_card_ref: str,
) -> tuple[
    list[SyntheticRateCardWorkbenchRow],
    list[SyntheticRateCardWorkbenchStateSummary],
    list[SyntheticRateCardWorkbenchCheck],
    int,
]:
    checks: list[SyntheticRateCardWorkbenchCheck] = []
    real_flags = _truthy_real_data_flags(card)
    checks.append(
        _check(
            "synthetic_candidate_declaration",
            card.get("status") == "candidate"
            and card.get("data_origin") == "synthetic"
            and card.get("candidate_only") is True
            and card.get("contains_real_firm_data") is False
            and not real_flags,
            "Rate card must declare candidate synthetic-only status and no real-data flags.",
            rate_card_ref,
        )
    )
    checks.append(
        _check(
            "non_authoritative_declaration",
            card.get("not_promoted_canon") is True
            and card.get("real_rate_import_allowed") is False,
            "Rate card must remain non-canonical and block real-rate import.",
            rate_card_ref,
        )
    )

    card_id = card.get("rate_card_id")
    version = card.get("version")
    checks.append(
        _check(
            "identity_and_version_present",
            isinstance(card_id, str)
            and bool(card_id.strip())
            and isinstance(version, str)
            and bool(version.strip()),
            "Rate card must have a stable ID and version.",
            rate_card_ref,
        )
    )

    titles = card.get("titles")
    valid_titles = (
        isinstance(titles, list)
        and bool(titles)
        and all(isinstance(title, str) and title.strip() for title in titles)
        and len(set(titles)) == len(titles)
    )
    checks.append(
        _check(
            "title_catalog_complete",
            valid_titles,
            "Titles must be a non-empty, unique list of synthetic staffing roles.",
            rate_card_ref,
        )
    )
    title_set = set(titles) if valid_titles else set()

    carriers = _as_mapping(card.get("carriers"))
    checks.append(
        _check(
            "carrier_catalog_present",
            bool(carriers),
            "At least one synthetic carrier schedule is required.",
            rate_card_ref,
        )
    )

    rows: list[SyntheticRateCardWorkbenchRow] = []
    override_count = 0
    schedule_state_sets: dict[str, set[str]] = {}
    carrier_shape_valid = True
    schedule_complete = True
    rates_valid = True
    override_valid = True
    duplicate_cells = False
    seen_cells: set[tuple[str, str, str]] = set()

    for carrier_id in sorted(carriers):
        spec = _as_mapping(carriers[carrier_id])
        carrier_name = spec.get("carrier_name")
        effective_date = spec.get("effective_date")
        aliases = spec.get("aliases")
        schedule = _as_mapping(spec.get("schedule"))
        if not (
            isinstance(carrier_name, str)
            and carrier_name.strip()
            and isinstance(effective_date, str)
            and effective_date.strip()
            and isinstance(aliases, list)
            and bool(aliases)
            and all(isinstance(alias, str) and alias.strip() for alias in aliases)
            and bool(schedule)
        ):
            carrier_shape_valid = False
        schedule_state_sets[str(carrier_id)] = set(schedule)
        for state in sorted(schedule):
            role_rates = _as_mapping(schedule[state])
            if set(role_rates) != title_set:
                schedule_complete = False
            for title in sorted(role_rates):
                rate = role_rates[title]
                try:
                    numeric_rate = float(rate)
                except (TypeError, ValueError):
                    rates_valid = False
                    continue
                if numeric_rate <= 0:
                    rates_valid = False
                    continue
                key = (str(carrier_id), str(state), str(title))
                if key in seen_cells:
                    duplicate_cells = True
                    continue
                seen_cells.add(key)
                rows.append(
                    SyntheticRateCardWorkbenchRow(
                        carrier_id=str(carrier_id),
                        carrier_name=str(carrier_name or carrier_id),
                        effective_date=str(effective_date or ""),
                        state=str(state),
                        title=str(title),
                        hourly_rate=numeric_rate,
                    )
                )

        overrides = _as_mapping(spec.get("named_timekeeper_overrides"))
        for _timekeeper_id, override in overrides.items():
            override_count += 1
            override_spec = _as_mapping(override)
            override_state = override_spec.get("state")
            override_title = override_spec.get("title")
            approved_rate = override_spec.get("approved_rate")
            try:
                numeric_override = float(approved_rate)
            except (TypeError, ValueError):
                numeric_override = 0.0
            if (
                override_state not in schedule
                or override_title not in title_set
                or numeric_override <= 0
            ):
                override_valid = False

    expected_states = set.union(*schedule_state_sets.values()) if schedule_state_sets else set()
    schedules_align = bool(schedule_state_sets) and all(
        states == expected_states for states in schedule_state_sets.values()
    )
    aliases = _as_mapping(card.get("jurisdiction_aliases"))
    declared_states = {str(state) for state in aliases.values()}
    default_state = card.get("default_state")
    if isinstance(default_state, str) and default_state.strip():
        declared_states.add(default_state)
    declared_states_covered = bool(expected_states) and declared_states <= expected_states
    checks.extend(
        [
            _check(
                "carrier_metadata_complete",
                carrier_shape_valid,
                "Every carrier needs a synthetic name, aliases, effective date, and schedule.",
                rate_card_ref,
            ),
            _check(
                "schedule_states_aligned",
                schedules_align,
                "Every carrier must provide the same state schedule coverage.",
                rate_card_ref,
            ),
            _check(
                "alias_and_default_states_covered",
                declared_states_covered,
                "Default and jurisdiction-alias states must be covered by every carrier schedule.",
                rate_card_ref,
            ),
            _check(
                "schedule_roles_complete",
                schedule_complete and bool(rows),
                "Every carrier/state schedule must provide exactly the declared titles.",
                rate_card_ref,
            ),
            _check(
                "positive_numeric_rates",
                rates_valid,
                "Every synthetic rate must be a positive number.",
                rate_card_ref,
            ),
            _check(
                "unique_rate_cells",
                not duplicate_cells,
                "Each carrier/state/title cell must be unique.",
                rate_card_ref,
            ),
            _check(
                "named_timekeeper_overrides_valid",
                override_valid,
                "Named synthetic overrides must reference a scheduled state/title and positive rate.",
                rate_card_ref,
            ),
        ]
    )

    summaries: list[SyntheticRateCardWorkbenchStateSummary] = []
    for carrier_id in sorted({row.carrier_id for row in rows}):
        for state in sorted({row.state for row in rows if row.carrier_id == carrier_id}):
            state_rows = [
                row for row in rows if row.carrier_id == carrier_id and row.state == state
            ]
            rates = [row.hourly_rate for row in state_rows]
            summaries.append(
                SyntheticRateCardWorkbenchStateSummary(
                    carrier_id=carrier_id,
                    carrier_name=state_rows[0].carrier_name,
                    state=state,
                    role_count=len(state_rows),
                    minimum_hourly_rate=min(rates),
                    maximum_hourly_rate=max(rates),
                    average_hourly_rate=sum(rates) / len(rates),
                )
            )
    return rows, summaries, checks, override_count


def build_synthetic_rate_card_workbench_report(
    rate_card_path: str | Path,
    *,
    repo_root: str | Path,
    generated_at: str | None = None,
) -> SyntheticRateCardWorkbenchReport:
    """Audit one synthetic card and assemble the local candidate workbench report."""

    repo_root_path = Path(repo_root)
    source_path = Path(rate_card_path)
    rate_card_ref = _relative_ref(source_path, repo_root_path)
    raw_text = source_path.read_text(encoding="utf-8")
    rate_card_sha256 = digest_text(raw_text)
    parsed = yaml.safe_load(raw_text)
    card = parsed if isinstance(parsed, dict) else {}
    rows, summaries, checks, override_count = _collect_rows_and_checks(
        card, rate_card_ref=rate_card_ref
    )
    failed_check_count = sum(check.status == "failed" for check in checks)
    payload_for_id = {
        "rate_card_sha256": rate_card_sha256,
        "methodology_version": METHODOLOGY_VERSION,
        "rows": [row.model_dump(mode="json") for row in rows],
        "checks": [check.model_dump(mode="json") for check in checks],
    }
    return SyntheticRateCardWorkbenchReport(
        synthetic_rate_card_workbench_report_id=(
            "synrateworkbench-" + digest_json(payload_for_id).removeprefix("sha256:")[:16]
        ),
        status=(
            "synthetic_rate_card_workbench_ready_for_review"
            if failed_check_count == 0
            else "blocked_by_synthetic_rate_card_workbench"
        ),
        rate_card_id=str(card.get("rate_card_id", "unknown")),
        rate_card_version=str(card.get("version", "unknown")),
        rate_card_ref=rate_card_ref,
        rate_card_sha256=rate_card_sha256,
        editable_source_ref=rate_card_ref,
        carrier_count=len({row.carrier_id for row in rows}),
        state_count=len({row.state for row in rows}),
        title_count=len({row.title for row in rows}),
        row_count=len(rows),
        named_timekeeper_override_count=override_count,
        rows=rows,
        state_summaries=summaries,
        checks=checks,
        failed_check_count=failed_check_count,
        workbook_filename=SYNTHETIC_RATE_CARD_WORKBOOK_FILENAME,
        markdown_filename=SYNTHETIC_RATE_CARD_WORKBENCH_MARKDOWN_FILENAME,
        display_banner={
            "summary": (
                "Synthetic candidate rate catalog only. It is not a carrier panel, firm "
                "rate card, benchmark, billing authority, or submission artifact."
            ),
            "candidate_only": True,
            "synthetic_only": True,
            "real_rate_import_allowed": False,
            "blocked_actions": [
                "real_rate_import",
                "matter_specific_rate_resolution_without_confirmation",
                "budget_submission",
                "carrier_submission",
                "exception_lake_write",
                "sqlite_write",
                "silent_learning",
            ],
        },
        candidate_exception_lake_labels=[
            "synthetic_rate_card_workbench_review_candidate",
            "rate_card_provenance_or_completeness_candidate",
        ],
        required_next_gates=[
            "human_review_before_synthetic_rate_card_change",
            "legal_knowledge_runtime_reviewed_snapshot_before_real_rate_replacement",
            "semantic_substrate_owner_review_before_canonical_promotion",
        ],
        generated_at=generated_at or _now_iso(),
    )


def _style_sheet(ws: Any, *, header_row: int) -> None:
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="234157")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"


def _write_sheet_banner(ws: Any) -> None:
    ws.append(
        [
            "Synthetic candidate-only local evidence. Not a real carrier or firm rate card; "
            "not billing authority or a submission artifact."
        ]
    )
    ws.append([])


def _write_workbook(
    report: SyntheticRateCardWorkbenchReport,
    path: Path,
) -> None:
    """Write a new macro-free workbook containing only the audited synthetic catalog."""

    from openpyxl import Workbook

    workbook = Workbook()
    read_me = workbook.active
    read_me.title = "Read Me"
    read_me.append([_safe_text("Synthetic Rate Card Workbench")])
    read_me.append([_safe_text("Status"), _safe_text(report.status)])
    read_me.append([_safe_text("Source"), _safe_text(report.editable_source_ref)])
    read_me.append([_safe_text("Source hash"), _safe_text(report.rate_card_sha256)])
    read_me.append(
        [_safe_text("Rate card"), _safe_text(f"{report.rate_card_id} v{report.rate_card_version}")]
    )
    read_me.append([_safe_text("Generated at"), _safe_text(report.generated_at)])
    read_me.append([_safe_text("Boundary"), _safe_text(report.display_banner["summary"])])
    read_me.append(["Editable source", "Edit the checked-in synthetic YAML, then regenerate."])
    read_me.append(["Blocked", ", ".join(report.display_banner["blocked_actions"])])
    read_me.column_dimensions["A"].width = 20
    read_me.column_dimensions["B"].width = 118

    rate_card = workbook.create_sheet("Rate Card")
    _write_sheet_banner(rate_card)
    rate_card.append(["Carrier ID", "Carrier", "Effective Date", "State", "Title", "Hourly Rate"])
    for row in report.rows:
        rate_card.append(
            [
                _safe_text(row.carrier_id),
                _safe_text(row.carrier_name),
                _safe_text(row.effective_date),
                _safe_text(row.state),
                _safe_text(row.title),
                row.hourly_rate,
            ]
        )
    _style_sheet(rate_card, header_row=3)
    for column, width in {"A": 25, "B": 28, "C": 16, "D": 10, "E": 22, "F": 16}.items():
        rate_card.column_dimensions[column].width = width
    for cell in rate_card["F"][1:]:
        cell.number_format = "$#,##0.00"

    summary = workbook.create_sheet("State Role Summary")
    _write_sheet_banner(summary)
    summary.append(["Carrier ID", "Carrier", "State", "Roles", "Minimum", "Maximum", "Average"])
    for row in report.state_summaries:
        summary.append(
            [
                _safe_text(row.carrier_id),
                _safe_text(row.carrier_name),
                _safe_text(row.state),
                row.role_count,
                row.minimum_hourly_rate,
                row.maximum_hourly_rate,
                row.average_hourly_rate,
            ]
        )
    _style_sheet(summary, header_row=3)
    for column, width in {"A": 25, "B": 28, "C": 10, "D": 10, "E": 14, "F": 14, "G": 14}.items():
        summary.column_dimensions[column].width = width
    for column in ("E", "F", "G"):
        for cell in summary[column][1:]:
            cell.number_format = "$#,##0.00"

    change_log = workbook.create_sheet("Change Log")
    _write_sheet_banner(change_log)
    change_log.append(["Control", "Value"])
    change_log.append(["Data origin", _safe_text(report.data_origin)])
    change_log.append(["Candidate only", str(report.candidate_only).lower()])
    change_log.append(["Non-authoritative", str(report.non_authoritative).lower()])
    change_log.append(["Real rate import allowed", str(report.real_rate_import_allowed).lower()])
    change_log.append(["External writes performed", str(report.external_writes_performed).lower()])
    change_log.append(["Lake write performed", str(report.lake_write_performed).lower()])
    change_log.append(["Workbook formulas", "none"])
    _style_sheet(change_log, header_row=3)
    change_log.column_dimensions["A"].width = 30
    change_log.column_dimensions["B"].width = 110

    workbook.properties.title = "Synthetic Rate Card Workbench"
    workbook.properties.subject = "Synthetic candidate-only local evidence"
    workbook.properties.creator = "LawFirm-os-intake"
    workbook.properties.lastModifiedBy = "LawFirm-os-intake"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_markdown(report: SyntheticRateCardWorkbenchReport, path: Path) -> None:
    lines = [
        "# Synthetic Rate Card Workbench",
        "",
        report.display_banner["summary"],
        "",
        f"- Source: `{report.editable_source_ref}`",
        f"- Source hash: `{report.rate_card_sha256}`",
        f"- Status: `{report.status}`",
        f"- Rows: `{report.row_count}`",
        "",
        "## Checks",
        "",
    ]
    lines.extend(
        f"- `{check.status}` `{check.check_id}`: {check.message}" for check in report.checks
    )
    lines.extend(["", "## Required Gates", ""])
    lines.extend(f"- {gate}" for gate in report.required_next_gates)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_synthetic_rate_card_workbench(
    *,
    repo_root: str | Path,
    out_dir: str | Path,
    generated_at: str | None = None,
) -> tuple[SyntheticRateCardWorkbenchReport, Path]:
    """Render local candidate artifacts from the fixed checked-in synthetic source."""

    root = Path(repo_root)
    source_path = root / "config" / "synthetic-carrier-rate-card.yaml"
    report = build_synthetic_rate_card_workbench_report(
        source_path,
        repo_root=root,
        generated_at=generated_at,
    )
    run_dir = Path(out_dir)
    run_dir.mkdir(parents=True, exist_ok=True)
    write_json(
        run_dir / SYNTHETIC_RATE_CARD_WORKBENCH_REPORT_FILENAME, report.model_dump(mode="json")
    )
    _write_markdown(report, run_dir / SYNTHETIC_RATE_CARD_WORKBENCH_MARKDOWN_FILENAME)
    if report.status == "synthetic_rate_card_workbench_ready_for_review":
        _write_workbook(report, run_dir / SYNTHETIC_RATE_CARD_WORKBOOK_FILENAME)
    return report, run_dir
